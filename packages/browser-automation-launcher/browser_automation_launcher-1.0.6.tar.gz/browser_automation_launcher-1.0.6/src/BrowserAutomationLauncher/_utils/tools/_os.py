"""
Copyright (c) 2025-now Martian Bugs All rights reserved.
Build Date: 2025-05-19
Author: Martian Bugs
Description: 系统工具类, 例如文件操作、路径操作等
"""

from os import path, remove
from typing import Any

from openpyxl import load_workbook


class OsTools:
    @staticmethod
    def file_remove(file_path: str):
        """
        删除文件

        Args:
            file_path: 文件路径
        Returns:
            是否删除成功, 如果文件不存在或不是文件, 则返回False
        """

        if not path.exists(file_path) or not path.isfile(file_path):
            return False

        remove(file_path)
        return True

    @staticmethod
    def xlsx_read(file_path: str, sheet_name: str = None, titles: list[str] = None):
        """
        读取 xlsx 文件内容

        Args:
            file_path: xlsx 文件路径
            sheet_name: 工作表名称，默认激活的工作表
            titles: 标题列表，如果不指定，则读取第一行作为标题
        Returns:
            xlsx 内容列表
        """

        wb = load_workbook(file_path)
        sheet = (
            wb.active
            if not sheet_name or not isinstance(sheet_name, str)
            else wb[sheet_name]
        )

        titles_dict = {}
        title_row = next(sheet.rows)
        if not titles or not isinstance(titles, list):
            for ci, col in enumerate(title_row):
                titles_dict[col.value] = ci
        else:
            titles_dict = {k: -1 for k in titles}
            for ci, col in enumerate(title_row):
                if col.value in titles_dict:
                    titles_dict[col.value] = ci

        records: list[dict[str, Any]] = []
        for row in sheet.iter_rows(min_row=2):
            record = {k: v for k, ci in titles_dict.items() for v in [row[ci].value]}
            records.append(record)

        return records
