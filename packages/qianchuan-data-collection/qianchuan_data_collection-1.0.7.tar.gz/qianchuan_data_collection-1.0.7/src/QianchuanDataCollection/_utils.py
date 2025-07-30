from copy import deepcopy
from datetime import datetime, timedelta
from os import path, remove
from typing import Any
from urllib.parse import urlparse

from openpyxl import load_workbook


class Utils:
    @staticmethod
    def date_yesterday(pattern='%Y-%m-%d'):
        """
        获取前一天的日期

        Args:
            pattern: 日期格式
        """

        return Utils.date_calculate(days=1, pattern=pattern)

    @staticmethod
    def date_calculate(days: int, pattern='%Y-%m-%d'):
        """
        日期计算

        Args:
            days: 日期偏移量, 负数表示向后推
            pattern: 日期格式
        """

        return (datetime.now() - timedelta(days=days)).strftime(pattern)

    @staticmethod
    def date_diff_days(a: str, b: str, pattern='%Y-%m-%d'):
        """
        计算两个日期间隔的天数
        - 正数表示 a 日期在 b 日期之后

        Args:
            a: 日期字符串
            b: 日期字符串
        """

        a_dt = datetime.strptime(a, pattern)
        b_dt = datetime.strptime(b, pattern)

        return (a_dt - b_dt).days

    @staticmethod
    def same_url(a: str, b: str):
        """
        检查两个 url 是否域名及路径是否一致

        Args:
            a: 第一个 url
            b: 第二个 url
        Returns:
            是否一致
        """

        a_result = urlparse(a)
        b_result = urlparse(b)

        is_same = a_result.netloc == b_result.netloc and a_result.path == b_result.path

        return is_same

    @staticmethod
    def dict_mapping(data: dict, dict_table: dict[str, str]):
        """
        字典表字段映射

        Args:
            data: 待映射的字典
            dict_table: 字典表
        """

        result = {}
        for text, key in dict_table.items():
            result[text] = data.get(key)

        return result

    @staticmethod
    def dict_format__float(data: dict, fields: list[str] = None, precision: int = 2):
        """
        将字典数据中的指定字段格式化为 float 类型

        Args:
            data: 待格式化的字典
            fields: 需要格式化的字段列表, 如果为 None, 则表示所有字段
            precision: 保留小数位数, 默认 2 位
        Returns:
            格式化后的字典
        """

        _fields = fields if fields and isinstance(fields, list) else data.keys()

        _data = deepcopy(data)
        for field in _fields:
            if field not in _data:
                continue

            value = _data[field]
            if not isinstance(value, (int, float)):
                continue

            _data[field] = value / 10**precision

        return _data

    @staticmethod
    def dict_format__round(data: dict, fields: list[str] = None, precision: int = 2):
        """
        将字典数据中的指定字段作四舍五入处理

        Args:
            data: 待格式化的字典
            fields: 需要格式化的字段列表, 如果为 None, 则表示所有字段
            precision: 保留小数位数, 默认 2 位
        Returns:
            格式化后的字典
        """

        _fields = fields if fields and isinstance(fields, list) else data.keys()

        _data = deepcopy(data)
        for field in _fields:
            if field not in _data:
                continue

            value = _data[field]
            if not isinstance(value, float):
                continue

            _data[field] = round(value, precision)

        return _data

    @staticmethod
    def dict_format__ratio(data: dict, fields: list[str] = None, ratio: int = 2):
        """
        将字典数据中的指定字段转为比率, 例如百分比/千分比等

        Args:
            data: 待格式化的字典
            fields: 需要格式化的字段列表, 如果为 None, 则表示所有字段
            ratio: 比率, 默认 2 及百分比
        Returns:
            格式化后的字典
        """

        _fields = fields if fields and isinstance(fields, list) else data.keys()

        _data = deepcopy(data)
        for field in _fields:
            if field not in _data:
                continue

            value = _data[field]
            if not isinstance(value, (int, float)):
                continue

            _data[field] = value * (10**ratio)

        return _data

    @staticmethod
    def dict_format__strip(
        data: dict,
        fields: list[str] = None,
        prefix: list[str] = None,
        suffix: list[str] = None,
    ):
        """
        格式化字典数据中的指定字段, 去除前后空格及指定前后缀

        Args:
            data: 待格式化的字典
            fields: 需要格式化的字段列表, 如果为 None, 则表示所有字段
            prefix: 需要去除的前缀列表
            suffix: 需要去除的后缀列表
        Returns:
            格式化后的字典
        """

        _fields = fields if fields and isinstance(fields, list) else data.keys()

        _data = deepcopy(data)
        for field in _fields:
            if field not in _data:
                continue

            value = _data[field]
            if not isinstance(value, str):
                continue

            value = value.strip()
            if prefix and isinstance(prefix, list):
                for c in prefix:
                    value = value.lstrip(c)

            if suffix and isinstance(suffix, list):
                for c in suffix:
                    value = value.rstrip(c)

            _data[field] = value

        return _data

    @staticmethod
    def dict_format__number(
        data: dict, fields: list[str] = None, exclude_fields: list[str] = None
    ):
        """
        格式化字典数据中的指定字段, 将字符串转为数字

        Args:
            data: 待格式化的字典
            fields: 需要格式化的字段列表, 如果为 None, 则表示所有字段
            exclude_fields: 排除的字段列表
        Returns:
            格式化后的字典
        """

        _fields = fields if fields and isinstance(fields, list) else data.keys()

        _data = deepcopy(data)
        for field in _fields:
            if field not in _data:
                continue

            value = _data[field]
            if not isinstance(value, str):
                continue

            try:
                value = value.replace(',', '')
                value = float(value) if '.' in value else int(value)
            except ValueError:
                continue

            _data[field] = value

        if exclude_fields and isinstance(exclude_fields, list):
            for field in exclude_fields:
                if field not in _data:
                    continue
                _data[field] = data[field]

        return _data

    @staticmethod
    def xlsx_read(file_path: str, sheet_name: str = None, titles: list[str] = None):
        """
        读取 xlsx 文件内容

        Args:
            file_path: xlsx 文件路径
            sheet_name: 工作表名称，默认激活的工作表
            titles: 标题列表，如果不指定，则读取第一行作为标题
        Returns:
            xlsx 内容列表
        """

        wb = load_workbook(file_path)
        sheet = (
            wb.active
            if not sheet_name or not isinstance(sheet_name, str)
            else wb[sheet_name]
        )

        titles_dict = {}
        title_row = next(sheet.rows)
        if not titles or not isinstance(titles, list):
            for ci, col in enumerate(title_row):
                titles_dict[col.value] = ci
        else:
            titles_dict = {k: -1 for k in titles}
            for ci, col in enumerate(title_row):
                if col.value in titles_dict:
                    titles_dict[col.value] = ci

        records: list[dict[str, Any]] = []
        for row in sheet.iter_rows(min_row=2):
            record = {k: v for k, ci in titles_dict.items() for v in [row[ci].value]}
            records.append(record)

        return records

    @staticmethod
    def file_remove(file_path: str):
        """
        删除文件

        Args:
            file_path: 文件路径
        Returns:
            是否删除成功
        """

        if not path.exists(file_path) or not path.isfile(file_path):
            return False

        remove(file_path)
        return True
