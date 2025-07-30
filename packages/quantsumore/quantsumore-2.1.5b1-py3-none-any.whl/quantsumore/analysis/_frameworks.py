# -*- coding: utf-8 -*-
#
## ╭────────────────────────────────────────────────────────────────────────────────────────────╮
## │  Library         : doydl's Finance API Client — quantsumore                                 │
## │                                                                                             │
## │                                                                                             │
## │  Description     : `quantsumore` is a comprehensive Python library designed to streamline   │
## │                    the process of accessing and analyzing real-time financial data across   │
## │                    various markets. It provides specialized API clients to fetch data       │
## │                    from multiple financial instruments, including:                          │
## │                      - Cryptocurrencies                                                     │
## │                      - Equities and Stock Markets                                           │
## │                      - Foreign Exchange (Forex)                                             │
## │                      - Treasury Instruments                                                 │
## │                      - Consumer Price Index (CPI) Metrics                                   │
## │                                                                                             │
## │                    The library offers a unified interface for retrieving diverse financial  │
## │                    data, enabling users to perform in-depth financial and technical         │
## │                    analysis. Whether you're developing trading algorithms, conducting       │
## │                    market research, or building financial dashboards, `quantsumore` serves  │
## │                    as a reliable and efficient tool in your data pipeline.                  │
## │                                                                                             │
## │                                                                                             │
## │  Key Features    : - Real-time data retrieval from multiple financial markets               │
## │                    - Support for various financial instruments and metrics                  │
## │                    - Simplified API clients for ease of integration                         │
## │                    - Designed for both personal and non-commercial use                      │
## │                                                                                             │
## │                                                                                             │
## │  Legal Disclaimer: `quantsumore` is an independent Python library and is not affiliated     │
## │                    with any financial institutions or data providers. Likewise, doydl       │
## │                    technologies is not affiliated with, endorsed by, or sponsored by any    │
## │                    government, corporate, or financial institutions. Users should verify    │
## │                    the accuracy of the data obtained and consult professional advice        │
## │                    before making investment decisions.                                      │
## │                                                                                             │
## │                                                                                             │
## │  Copyright       : © 2023–2025 by doydl technologies. All rights reserved.                  │
## │                                                                                             │
## │                                                                                             │
## │  License         : Licensed under the Apache License, Version 2.0 (the "License");          │
## │                    you may not use this file except in compliance with the License.         │
## │                    You may obtain a copy of the License at:                                 │
## │                                                                                             │
## │                        http://www.apache.org/licenses/LICENSE-2.0                           │
## │                                                                                             │
## │                    Unless required by applicable law or agreed to in writing, software      │
## │                    distributed under the License is distributed on an "AS IS" BASIS,        │
## │                    WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or          │
## │                    implied. See the License for the specific language governing             │
## │                    permissions and limitations under the License.                           │
## ╰────────────────────────────────────────────────────────────────────────────────────────────╯
#



import os
import pandas as pd
import datetime
import random

#────────── Third-party library imports (from PyPI or other package sources) ─────────────────────────────────
import openpyxl
from openpyxl.styles import NamedStyle, Font, Border, Side, Alignment, PatternFill
from openpyxl.utils import get_column_letter


# ━━━━━━━━━━━━━━ Core Module Implementation ━━━━━━━━━━━━━━━━━━━━━━━━━━
# This segment delineates the functional backbone of the module.
# It comprises the abstractions and behaviors essential for runtime
# execution—if applicable—encapsulated in class and function constructs.
# In minimal implementations, this may simply define constants, metadata,
# or serve as an interface placeholder.
statement_layouts = {
   "Income Statement":{
      "indentation_levels":{
         "Total Revenue":0,
         "Cost of Revenue":1,
         "Gross Profit":0,
         "Operating Expenses":0,
         "Research and Development":1,
         "Sales, General and Admin.":1,
         "Non-Recurring Items":1,
         "Other Operating Items":1,
         "Operating Income":0,
         "Add\\'l income/expense items":1,
         "Earnings Before Interest and Tax":0,
         "Interest Expense":1,
         "Earnings Before Tax":0,
         "Income Tax":1,
         "Minority Interest":1,
         "Equity Earnings/Loss Unconsolidated Subsidiary":1,
         "Net Income-Cont. Operations":0,
         "Net Income":0,
         # "Net Income Applicable to Common Shareholders":1
         "Net Income Applicable to Common Shareholders":0         
      },
      "parent_accounts":[
         "Total Revenue",
         "Operating Expenses"
      ],
      "subtotal_total_accounts":[
         "Gross Profit",
         "Operating Income",
         "Earnings Before Interest and Tax",
         "Earnings Before Tax",
         "Net Income-Cont. Operations",
         "Net Income",
         "Net Income Applicable to Common Shareholders"
      ]
   },
   "Balance Sheet":{
      "indentation_levels":{
         "Current Assets":0,
         "Cash and Cash Equivalents":1,
         "Short-Term Investments":1,
         "Net Receivables":1,
         "Inventory":1,
         "Other Current Assets":1,
         "Total Current Assets":0,
         "Long-Term Assets":0,
         "Long-Term Investments":1,
         "Fixed Assets":1,
         "Goodwill":1,
         "Intangible Assets":1,
         "Other Assets":1,
         "Deferred Asset Charges":1,
         "Total Assets":0,
         "Current Liabilities":0,
         "Accounts Payable":1,
         "Short-Term Debt / Current Portion of Long-Term Debt":1,
         "Other Current Liabilities":1,
         "Total Current Liabilities":0,
         "Long-Term Debt":0,
         "Other Liabilities":0,
         "Deferred Liability Charges":1,
         "Misc. Stocks":1,
         "Minority Interest":1,
         "Total Liabilities":0,
         "Stock Holders Equity":0,
         "Common Stocks":1,
         "Capital Surplus":1,
         "Retained Earnings":1,
         "Treasury Stock":1,
         "Other Equity":1,
         "Total Equity":0,
         "Total Liabilities & Equity":0
      },
      "parent_accounts":[
         "Current Assets",
         "Long-Term Assets",
         "Current Liabilities",
         "Stock Holders Equity"
      ],
      "subtotal_total_accounts":[
         "Total Current Assets",
         "Total Assets",
         "Total Current Liabilities",
         "Total Liabilities",
         "Total Equity",
         "Total Liabilities & Equity"
      ]
   },
   "Cash Flow Statement":{
      "indentation_levels":{
         "Net Income":0,
         "Cash Flows-Operating Activities":0,
         "Depreciation":1,
         "Net Income Adjustments":1,
         "Changes in Operating Activities":0,
         "Accounts Receivable":1,
         "Changes in Inventories":1,
         "Other Operating Activities":1,
         "Liabilities":1,
         "Net Cash Flow-Operating":0,
         "Cash Flows-Investing Activities":0,
         "Capital Expenditures":1,
         "Investments":1,
         "Other Investing Activities":1,
         "Net Cash Flows-Investing":0,
         "Cash Flows-Financing Activities":0,
         "Sale and Purchase of Stock":1,
         "Net Borrowings":1,
         "Other Financing Activities":1,
         "Net Cash Flows-Financing":0,
         "Effect of Exchange Rate":0,
         "Net Cash Flow":0
      },
      "parent_accounts":[
         "Cash Flows-Operating Activities",
         "Changes in Operating Activities",
         "Cash Flows-Investing Activities",
         "Cash Flows-Financing Activities"
      ],
      "subtotal_total_accounts":[
         "Net Cash Flow-Operating",
         "Net Cash Flows-Investing",
         "Net Cash Flows-Financing",
         "Net Cash Flow"
      ]
   }
} 




class WriteExcel:
    """
    This class provides a set of functionalities for writing financial statments to an Excel file, adjusting column widths
    automatically based on the data content, and applying various styles to the data to enhance readability and presentation.
    """
    def __init__(self, path=None):
        """Initialize the workbook, loading from path if provided."""    	
        self.path = path
        if path and os.path.exists(path):
            self.wb = openpyxl.load_workbook(path)
        else:
            self.wb = openpyxl.Workbook()
            if 'Sheet' in self.wb.sheetnames: # Remove the default 'Sheet' if it exists
                del self.wb['Sheet']
        self.base_width = 1.0  # Adjust as necessary
        self.min_width = 10  # Adjust based on observations

    def __enter__(self):
        return self

    def __exit__(self, exc_type, exc_val, exc_tb):
        self.wb.close()

    def __which_statement(self, financial_statement):
        accounts = list(financial_statement.index.values)
        if "Research and Development" in accounts:
            return "Income Statement"
        elif "Goodwill" in accounts:
            return "Balance Sheet"
        elif "Depreciation" in accounts:
            return "Cash Flow Statement"
        else:
            return None

    def __coerce_numeric_min(self, n):
        """ Converts the input to an integer or float and ensures a minimum value of 1."""                
        if not isinstance(n, str):
            n = str(n)
        try:
            if '.' in n:
                float_val = float(n)
                if float_val.is_integer():
                    result = int(float_val)
                else:
                    result = float_val
            else:
                result = int(n)
            return max(result, 1)
        except ValueError:
            return 11
           
    def __autofit_column_width(self, ws, df):
        """Adjust column widths based on the content of the DataFrame."""    	
        for col_idx, column in enumerate(df.columns, start=1):
            max_length = max(len(column), max(len(str(cell)) for cell in df[column]))
            adjusted_width = max(max_length * self.base_width, self.min_width)
            ws.column_dimensions[get_column_letter(col_idx)].width = adjusted_width
                
    def __apply_styles(self, sheet_name, font_style="Calibri Light", font_size=10):
        """Apply pre-defined styles for data, headers, numbers, dates, and bold formatting."""
        font_size = self.__coerce_numeric_min(n=font_size)
        styles = [
            NamedStyle(name=f"data_font_style_{sheet_name}", font=Font(bold=False, name=font_style, size=font_size), number_format="@"),
            NamedStyle(name=f"header_font_style_{sheet_name}", font=Font(bold=True, name=font_style, size=font_size), number_format="@"),
            NamedStyle(name=f"number_format_style_{sheet_name}", font=Font(name=font_style, size=font_size), number_format="_(* #,##0.00_);_(* (#,##0.00);_(* \"-\"??_);_(@_)"),
            NamedStyle(name=f"bold_font_style_{sheet_name}", font=Font(bold=True, name=font_style, size=font_size), number_format="@"),
            NamedStyle(name=f"bold_number_format_style_{sheet_name}", font=Font(bold=True, name=font_style, size=font_size), number_format="_(* #,##0.00_);_(* (#,##0.00);_(* \"-\"??_);_(@_)")
        ]
        for style in styles:
            if style.name not in self.wb.named_styles:
                self.wb.add_named_style(style)
                
    def __create_or_replace_sheet(self, sheet_name, overwrite=True):
        """
        Creates a new sheet in the workbook. If a sheet with the same name exists, it can either
        overwrite it (delete and create new) or create a new sheet with a unique name.
        
        Parameters:
        - sheet_name (str): Name of the sheet to create.
        - overwrite (bool): If True, deletes the existing sheet with the same name.
                            If False, a new sheet with an appended suffix (like '_1') is created.
        """
        if sheet_name in self.wb.sheetnames:
            if overwrite:
                del self.wb[sheet_name]
            else:
                suffix = 1
                new_sheet_name = f"{sheet_name}_{suffix}"
                while new_sheet_name in self.wb.sheetnames:
                    suffix += 1
                    new_sheet_name = f"{sheet_name}_{suffix}"
                sheet_name = new_sheet_name
        ws = self.wb.create_sheet(title=sheet_name)
        return ws
        
    # Write Financial Statement
    #-----------------------------------------------------------------------------------------------------------------------------------        
    def write_statement(self, df, reporting_structure=statement_layouts):
        """ Writes the financial statement to an Excel sheet with indentation, subtotal/total styling, and right alignment except the first column. """
        sheet_name = self.__which_statement(df)
        ws = self.__create_or_replace_sheet(sheet_name, overwrite=True)
        self.__apply_styles(sheet_name)

        right_align_style = Alignment(horizontal="right")

        light_grey_fill = PatternFill(start_color='DADADA', end_color='DADADA', fill_type='solid')
        
        index_name = df.index.name if df.index.name is not None else "Ending:"
        a1_cell = ws.cell(row=1, column=1, value=index_name)
        a1_cell.style = f"header_font_style_{sheet_name}"
        a1_cell.fill = light_grey_fill
        
        indentation_levels = reporting_structure[sheet_name]["indentation_levels"]
        parent_accounts = reporting_structure[sheet_name]["parent_accounts"]
        subtotal_total_accounts = reporting_structure[sheet_name]["subtotal_total_accounts"]

        for col_idx, column_name in enumerate(df.columns, start=2):
            header_cell = ws.cell(row=1, column=col_idx, value=column_name)
            header_cell.style = f"header_font_style_{sheet_name}"
            header_cell.alignment = right_align_style

        for row_idx, (index, row_data) in enumerate(df.iterrows(), start=2):
            indent = indentation_levels.get(index.strip(), 0)
            row_name = f"{'    ' * indent}{index}"

            is_bold = index in parent_accounts or index in subtotal_total_accounts
            row_style_name = f"bold_font_style_{sheet_name}" if is_bold else f"data_font_style_{sheet_name}"
            name_cell = ws.cell(row=row_idx, column=1, value=row_name)
            name_cell.style = row_style_name

            for col_idx, value in enumerate(row_data, start=2):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.style = f"bold_number_format_style_{sheet_name}" if is_bold else f"number_format_style_{sheet_name}"
                cell.alignment = right_align_style

                if index in subtotal_total_accounts[:-1]:
                    cell.border = Border(bottom=Side(style="thin"))
                elif index in subtotal_total_accounts[-1]:
                    cell.border = Border(bottom=Side(style="double"))

        self.__autofit_column_width(ws, df)

    
    def save(self, filename=None, overwrite=True):
        """Save the workbook to the specified filename, with optional overwrite."""    	
        try:
            if not self.wb.sheetnames:
                self.wb.create_sheet('Sheet1')
            if filename is None:
                filename = f'output_{datetime.datetime.now().strftime("%Y-%m-%d %H_%M_%S")}_{random.randint(1000, 5000)}.xlsx'
            else:
                if not filename.endswith(".xlsx"):
                    filename += ".xlsx"
            if not overwrite and os.path.exists(filename):
                raise FileExistsError(f"File '{filename}' already exists and 'overwrite' is set to False.")
            self.wb.save(filename)
            print(f"File '{filename}' has been saved successfully.")
            return True         
        except Exception as e:
            print(f"Error occurred while saving file: {e}")
            raise

    def close(self):
        """Close the workbook, releasing any associated resources."""    	
        self.wb.close()




def __dir__():
    return ['WriteExcel']

__all__ = ['WriteExcel']
