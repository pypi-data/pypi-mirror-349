import os

import pandas as pd
from openpyxl import load_workbook
from xlrd import open_workbook


# TODO: move to Excel Facade
def sheet_summariser(
    path,
    file_path_and_name,
    file_extension,
):
    # remove_empty_rows
    def remove_empty_rows(sheet):
        for row in sheet:
            remove_empty_cells(
                sheet,
                row,
            )

            # remove_empty_cells

    def remove_empty_cells(sheet, row):
        for cell in row:
            if cell.value != None:
                return

        sheet.delete_rows(row[0].row, 1)

        # get_excel_columns

    def get_excel_columns(sheet_data):
        list_with_values = []

        for cell in sheet_data[1]:
            list_with_values.append(
                cell.value,
            )

        return list_with_values

    # csv_sheet_summariser
    def csv_sheet_summariser(file_name):
        csv_data = pd.read_csv(
            os.path.join(
                path,
                file_name,
            ),
            encoding="latin-1",
        )

        csv_data = csv_data.dropna(
            how="all",
        ).reset_index(drop=True)

        dim = csv_data.shape

        rows = dim[0]
        cols = dim[1]
        dict_df = {
            "number_of_columns": [cols],
            "number_of_rows": [rows],
        }
        return pd.DataFrame(dict_df)

    # excel_sheet_summariser
    def excel_sheet_summariser(
        sheet_name,
        workbook,
        file_extension,
    ) -> pd.DataFrame:
        sheet = workbook[
            f"{sheet_name}"
        ]

        remove_empty_rows(sheet)

        if (
            file_extension == ".xlsx"
            or file_extension == ".xlsm"
        ):
            sheet_data = sheet.values
            sheet_rows = sheet.max_row
        else:
            sheet_rows = sheet.nrows
            sheet_data = (
                workbook.sheet_by_name(
                    sheet_name,
                )
            )

        if sheet_rows > 0:
            column_list = (
                get_excel_columns(sheet)
            )

            sheet_data_dataframe = (
                pd.DataFrame(
                    sheet_data,
                    columns=column_list,
                )
            )

            sheet_data_dataframe = sheet_data_dataframe.dropna(
                how="all",
            ).reset_index(
                drop=True,
            )

            dim = (
                sheet_data_dataframe.shape
            )

            number_of_rows = dim[0]

            number_of_columns = dim[1]

            excel_sheet_summary_dataframe = pd.DataFrame(
                {
                    "number_of_columns": [
                        number_of_columns,
                    ],
                    "number_of_rows": [
                        number_of_rows,
                    ],
                },
            )

            print(
                f"\nfound sheet {sheet_name}: \nsummary:\n{excel_sheet_summary_dataframe}\n",
            )

            print(
                f"found columns: {column_list}\n",
            )
        else:
            excel_sheet_summary_dataframe = pd.DataFrame(
                {
                    "number_of_columns": [
                        0,
                    ],
                    "number_of_rows": [
                        0,
                    ],
                },
            )

        return excel_sheet_summary_dataframe

    # main code
    sheet_summary_df = pd.DataFrame()

    if file_extension == ".csv":
        csv_sheet_summary_df = (
            csv_sheet_summariser(
                file_path_and_name,
            )
        )

        sheet_summary_df = (
            csv_sheet_summary_df
        )
        sheet_summary_df[
            "sheet_names"
        ] = os.path.basename(
            file_path_and_name,
        )

    else:
        if (
            file_extension == ".xlsx"
            or file_extension == ".xlsm"
        ):
            workbook = load_workbook(
                os.path.join(
                    path,
                    file_path_and_name,
                ),
            )

            sheet_names_list = (
                workbook.sheetnames
            )

            number_of_sheets = len(
                sheet_names_list,
            )

        else:
            workbook = open_workbook(
                os.path.join(
                    path,
                    file_path_and_name,
                ),
            )

            sheet_names_list = (
                workbook.sheet_names()
            )

            number_of_sheets = (
                workbook.nsheets
            )

        for (
            sheet_name
        ) in sheet_names_list:
            excel_sheet_summary = (
                excel_sheet_summariser(
                    sheet_name,
                    workbook,
                    file_extension,
                )
            )

            excel_sheet_summary[
                "sheet_names"
            ] = sheet_name.title()
            excel_sheet_summary[
                "number_of_sheets"
            ] = number_of_sheets

            sheet_summary_df = pd.concat(
                [
                    sheet_summary_df,
                    excel_sheet_summary,
                ],
            )

    return sheet_summary_df


# get_sheet_summary_from_directory
def get_sheet_summary_from_directory(
    summary_dataframe,
    parent_directory_path,
    files,
    valid_file_extensions,
):
    for file_name in files:
        _, file_extension = (
            os.path.splitext(file_name)
        )

        if (
            file_extension
            in valid_file_extensions
        ):
            print(
                f"*********summarising file {file_name} in {parent_directory_path}**********\n",
            )

            sheet_summary = sheet_summariser(
                parent_directory_path,
                file_name,
                file_extension,
            )

            sheet_summary[
                "file_names"
            ] = file_name

            sheet_summary[
                "parent_directory_paths"
            ] = parent_directory_path

            summary_dataframe = pd.concat(
                [
                    summary_dataframe,
                    sheet_summary,
                ],
            )

    return summary_dataframe


# summarise_sheet_files_in_directory
def summarise_sheet_files_in_directory(
    directory_path_and_name,
    valid_file_extensions,
) -> pd.DataFrame:
    print(
        f"\n------reading directory {directory_path_and_name}---------\n",
    )

    sheet_summary_report_schema = [
        "parent_directory_paths",
        "file_names",
        "number_of_sheets",
        "sheet_names",
        "number_of_columns",
        "number_of_rows",
    ]

    sheet_summary_report_dataframe = pd.DataFrame(
        columns=sheet_summary_report_schema,
    )

    for (
        parent_directory_path,
        sub_directories,
        files,
    ) in os.walk(
        directory_path_and_name,
    ):
        sheet_summary_report_dataframe = get_sheet_summary_from_directory(
            sheet_summary_report_dataframe,
            parent_directory_path,
            files,
            valid_file_extensions,
        )

    return (
        sheet_summary_report_dataframe
    )
