#PowerDB VERSION 2.2.3.2
#Created solely by WeDu, published in 5/20/25
import re
import stat
import os
import sys
import openpyxl
#requirements: standard modules + openpyxl(*1)
# *1 : third party library(for exporting .pdb tables into Excel sheet)
class InnerFunctions:
    def __init__(self):
        pass
    def get_the_word_inbetween(self, text, start_char, end_char):
        if not isinstance(text, str) or not isinstance(start_char, str) or not isinstance(end_char, str):
            raise TypeError("Inputs must be strings.")
        if not start_char or not end_char:
            raise ValueError("start_char and end_char cannot be empty.")
        if start_char not in text or end_char not in text:
            return None
        start_index = text.find(start_char)
        end_index = text.find(end_char, start_index + 1)
        if end_index <= start_index:
            return None
        return text[start_index + 1:end_index]
    def count_occurrences(self, word, string):
        if not isinstance(word, str) or not isinstance(string, str):
            raise TypeError("Inputs must be strings.")
        if not word:
            raise ValueError("word cannot be empty.")
        count = 0
        word_len = len(word)
        string_len = len(string)
        if word_len > string_len:
            return 0
        for i in range(string_len - word_len + 1):
            if string[i:i + word_len] == word:
                count += 1
        return count
    def get_line_of_phrase_in_text(self, text, phrase):
        if not isinstance(text, str) or not isinstance(phrase, str):
            raise TypeError("Inputs must be strings.")
        if not phrase:
            raise ValueError("phrase cannot be empty.")
        text = text.replace('\r\n', '\n').replace('\r', '\n') # Normalize line endings
        lines = text.splitlines()
        for line in lines:
            if phrase in line:
                return line.replace(phrase, "").strip()
        return None
    def modify_line_containing_word(self, text, word, new_line_content):
        if not isinstance(text, str) or not isinstance(word, str) or not isinstance(new_line_content, str):
            raise TypeError("Inputs must be strings.")
        if not word:
            raise ValueError("word cannot be empty.")
        text = text.replace('\r\n', '\n').replace('\r', '\n') # Normalize line endings
        lines = text.splitlines()
        for i, line in enumerate(lines):
            if word in line:
                lines[i] = new_line_content
                return os.linesep.join(lines) #os.linesep here is ok.
        return text
    def group_by_element(self, input_list, index):
        if not isinstance(input_list, list):
            raise TypeError("input_list must be a list.")
        if not isinstance(index, int):
            raise TypeError("index must be an integer.")
        grouped_list = {}
        for sublist in input_list:
            if not isinstance(sublist, list):
                raise ValueError("Inner items must be lists.")
            if not sublist:
                raise ValueError("Sublist cannot be empty")
            if index >= len(sublist):
                raise ValueError("Index is out of range for a sublist.")
            element = sublist[index]
            if element not in grouped_list:
                grouped_list[element] = []
            grouped_list[element].append(sublist)
        return list(grouped_list.values())
    def add_data_to_inner_lists(self, main_list, second_list):
        if not isinstance(main_list, list) or not isinstance(second_list, list):
            raise TypeError("Inputs must be lists.")
        result = []
        for i, inner_item in enumerate(main_list):
            if not isinstance(inner_item, list):
                raise ValueError("Inner items of main_list must be lists.")
            if i < len(second_list):
                result.append(inner_item + [second_list[i]])
            else:
                result.append(inner_item + [None])
                print("Warning: second_list is shorter than expected. Filling with None.", file=sys.stderr)
        return result
    def combine_lists(self, input_list):
        if not isinstance(input_list, list):
            raise TypeError("Input must be a list.")
        output_list = []
        for inner_list in input_list:
            if not isinstance(inner_list, list):
                raise ValueError("Inner items must be lists.")
            output_list.extend(inner_list)
        return output_list
inner_functions = InnerFunctions()
class CreateOperations:
    def __init__(self):
        pass
    def _normalize_path(self, file_path):
        if not isinstance(file_path, str):
            raise TypeError("file_path must be a string.")
        filepath = os.path.abspath(os.path.normpath(file_path))
        return filepath
    def _read_file_bytes(self, file_path):
        if not isinstance(file_path, str):
            raise TypeError("file_path must be a string.")
        filepath = self._normalize_path(file_path)
        if not os.path.exists(filepath):
            raise FileNotFoundError(f"File '{filepath}' does not exist.")
        try:
            with open(filepath, 'rb') as f:
                return f.read()
        except OSError as e:
            raise OSError(f"OS Error while reading file '{filepath}': {e}")
    def _write_file_bytes(self, file_path, data):
        if not isinstance(file_path, str):
            raise TypeError("file_path must be a string.")
        if not isinstance(data, bytes):
            raise TypeError("data must be bytes.")
        filepath = self._normalize_path(file_path)
        try:
            with open(filepath, 'wb') as f:
                f.write(data)
        except OSError as e:
            raise OSError(f"OS Error while writing file '{filepath}': {e}")
    def _append_file_bytes(self, file_path, data):
        if not isinstance(file_path, str):
            raise TypeError("file_path must be a string.")
        if not isinstance(data, bytes):
            raise TypeError("data must be bytes.")
        filepath = self._normalize_path(file_path)
        try:
            with open(filepath, 'ab') as f:
                f.write(data)
        except OSError as e:
            raise OSError(f"OS Error while appending file '{filepath}': {e}")
    def make_db(self, file_path):
        if not isinstance(file_path, str):
            raise TypeError("file_path must be a string.")
        filepath = self._normalize_path(file_path)
        if not filepath.lower().endswith('.pdb'):
            filepath = f'{filepath}.pdb'
        if os.path.exists(filepath):
            print(f"Error: File '{filepath}' already exists.")
            return
        try:
            with open(filepath, 'wb') as makeDBX:
                makeDBX.write(b'#POWER_DB')
        except OSError as e:
            raise OSError(f"OS Error while creating file '{filepath}': {e}")
    def make_container(self, file_path, name):
        if not isinstance(file_path, str):
            raise TypeError("file_path must be a string.")
        if not isinstance(name, str):
            raise TypeError("name must be a string.")
        if not name:
            raise ValueError("name cannot be empty.")
        filepath = self._normalize_path(file_path)
        if not os.path.exists(filepath):
            print(f"File '{filepath}' not found.")
            return
        try:
            file_content_bytes = self._read_file_bytes(filepath)
            file_content_str = file_content_bytes.decode('utf-8', errors='surrogateescape')
            # Normalize line endings
            file_content_str = file_content_str.replace('\r\n', '\n')
            num = inner_functions.count_occurrences('$<', file_content_str)
            container_string = f"{os.linesep if not file_content_str.endswith('\n') else ''}$<{num},{name}>"
            container_bytes = container_string.encode('utf-8')
            if f'$<{num},{name}>' not in file_content_str:
                self._append_file_bytes(filepath, container_bytes)
            else:
                print(f"Container '{name}' already exists.")
        except OSError as e:
            print(f"OS Error: {e}")
    def make_table(self, file_path, name):
        if not isinstance(file_path, str):
            raise TypeError("file_path must be a string.")
        if not isinstance(name, str):
            raise TypeError("name must be a string.")
        if not name:
            raise ValueError("name cannot be empty.")
        filepath = self._normalize_path(file_path)
        if not os.path.exists(filepath):
            print(f"File '{filepath}' not found.")
            return
        try:
            file_content_bytes = self._read_file_bytes(filepath)
            file_content_str = file_content_bytes.decode('utf-8', errors='surrogateescape')
            file_content_str = file_content_str.replace('\r\n', '\n')
            num = inner_functions.count_occurrences('&<', file_content_str)
            table_string = f"{os.linesep if not file_content_str.endswith('\n') else ''}&<{num}^{name}>"
            table_bytes = table_string.encode('utf-8')
            if f'&<{num}^{name}>' not in file_content_str:
                self._append_file_bytes(filepath, table_bytes)
            else:
                print(f"Table '{name}' already exists.")
        except OSError as e:
            print(f"OS Error: {e}")
    def set_file_permissions(self, file_path, mode=0o660):
        if not isinstance(file_path, str):
            raise TypeError("file_path must be a string.")
        if not isinstance(mode, int):
            raise TypeError("mode must be an integer.")
        filepath = self._normalize_path(file_path)
        if not os.path.exists(filepath):
            print(f"File '{filepath}' not found.")
            return
        try:
            os.chmod(filepath, mode)
        except OSError as e:
            raise OSError(f"OS Error while setting permissions for '{filepath}': {e}")
    def check_file_permissions(self, file_path, required_mode=stat.S_IRUSR | stat.S_IWUSR):
        if not isinstance(file_path, str):
            raise TypeError("file_path must be a string.")
        if not isinstance(required_mode, int):
            raise TypeError("required_mode must be an integer.")
        filepath = self._normalize_path(file_path)
        if not os.path.exists(filepath):
            print(f"File '{filepath}' not found.")
            return False
        try:
            mode = stat.S_IMODE(os.stat(filepath).st_mode)
            return bool(mode & required_mode)
        except OSError as e:
            raise OSError(f"OS Error while checking permissions for '{filepath}': {e}")
create = CreateOperations()
class container_data_class:
    def __init__(self):
        pass
    def _normalize_path(self, file_path):
        if not isinstance(file_path, str):
            raise TypeError("file_path must be a string.")
        filepath = os.path.abspath(os.path.normpath(file_path))
        if not os.path.exists(filepath):
            raise FileNotFoundError(f"File '{filepath}' does not exist.")
        return filepath
    def _read_file_bytes(self, file_path):
        filepath = self._normalize_path(file_path)
        if not self._check_file_permissions(filepath, stat.S_IRUSR):
            raise PermissionError(f"Insufficient permissions to read file: '{filepath}'")
        try:
            with open(filepath, 'rb') as f:
                return f.read()
        except OSError as e:
            raise OSError(f"OS Error while reading file '{filepath}': {e}")
    def _write_file_bytes(self, file_path, data):
        filepath = self._normalize_path(file_path)
        if not isinstance(data, bytes):
            raise TypeError("data must be bytes.")
        if not self._check_file_permissions(filepath, stat.S_IWUSR):
            self._set_file_permissions(filepath)
            if not self._check_file_permissions(filepath, stat.S_IWUSR):
                raise PermissionError(f"Insufficient permissions to write to file: '{filepath}'")
        try:
            with open(filepath, 'wb') as f:
                f.write(data)
        except OSError as e:
            raise OSError(f"OS Error while writing file '{filepath}': {e}")
    def _append_file_bytes(self, file_path, data):
        filepath = self._normalize_path(file_path)
        if not isinstance(data, bytes):
            raise TypeError("data must be bytes.")
        if not self._check_file_permissions(filepath, stat.S_IWUSR):
            self._set_file_permissions(filepath)
            if not self._check_file_permissions(filepath, stat.S_IWUSR):
                raise PermissionError(f"Insufficient permissions to append to file: '{filepath}'")
        try:
            with open(filepath, 'ab') as f:
                f.write(data)
        except OSError as e:
            raise OSError(f"OS Error while appending file '{filepath}': {e}")
    def _check_file_permissions(self, file_path, required_mode=stat.S_IRUSR | stat.S_IWUSR):
        if not isinstance(file_path, str):
            raise TypeError("file_path must be a string.")
        if not isinstance(required_mode, int):
            raise TypeError("required_mode must be an integer.")
        filepath = self._normalize_path(file_path)
        try:
            mode = stat.S_IMODE(os.stat(filepath).st_mode)
            return bool(mode & required_mode)
        except OSError as e:
            raise OSError(f"OS Error while checking permissions for '{filepath}': {e}")
    def _set_file_permissions(self, file_path, mode=0o660):
        if not isinstance(file_path, str):
            raise TypeError("file_path must be a string.")
        if not isinstance(mode, int):
            raise TypeError("mode must be an integer.")
        filepath = self._normalize_path(file_path)
        try:
            os.chmod(filepath, mode)
        except OSError as e:
            raise OSError(f"OS Error while setting permissions for '{filepath}': {e}")
    def getname(self, file, id):
        if not isinstance(file, str):
            raise TypeError("file must be a string")
        if not isinstance(id, int):
            raise TypeError("id must be an int")
        if id < 0:
            raise ValueError("id must be >= 0")
        try:
            r = self._read_file_bytes(file).decode('utf-8', errors='surrogateescape')
            pattern = rf'\$<({id}),([^>]*)>'
            match = re.search(pattern, r)
            if match:
                return match.group(2)
            else:
                return None
        except FileNotFoundError:
            return None
        except OSError:
            return None
    def getid(self, file, name, plogic=True):
        if not isinstance(file, str):
            raise TypeError("file must be a string.")
        if not isinstance(name, str):
            raise TypeError("name must be a string.")
        if not isinstance(plogic, bool):
            raise TypeError("plogic must be a bool.")
        if not name:
            raise ValueError("name cannot be empty")
        try:
            r = self._read_file_bytes(file).decode('utf-8', errors='surrogateescape')
            pattern = rf'\$<(\d+),{re.escape(name)}>'
            match = re.search(pattern, r)
            if match:
                return int(match.group(1)) if plogic else int(match.group(1)) + 1
            else:
                return -1 if plogic else 0
        except FileNotFoundError:
            return -1 if plogic else 0
        except OSError:
            return -1 if plogic else 0
    def insert(self, file, data, address=None, showrelational=False):
        if not isinstance(file, str):
            raise TypeError("file must be a string.")
        if not isinstance(data, str):
            raise TypeError("data must be a string.")
        if address is not None and not isinstance(address, list):
            raise TypeError("address must be a list or none")
        if not isinstance(showrelational, bool):
            raise TypeError("showrelational must be a bool")
        if address is None:
            address = []
        if len(address) != 2:
            raise ValueError("address must contain containerid and sectorid")
        containerid, sectorid = address
        if not isinstance(containerid, int) or not isinstance(sectorid, int):
            raise TypeError("containerid and sectorid must be integers")
        if containerid < 0 or sectorid < 0:
            raise ValueError("containerid and sectorid must be >= 0")
        try:
            r = self._read_file_bytes(file).decode('utf-8', errors='surrogateescape')
            num_sectors = self.numbersectors(file, containerid)
            if showrelational:
                print(sectorid, num_sectors)
            if not self.check(file, 'sector', [containerid, sectorid]):
                if sectorid - num_sectors <= 1:
                    new_data = f"!<[{containerid},{sectorid}],{data}>!".encode('utf-8', errors='surrogateescape')
                    if r.endswith('\n'): # Changed from os.linesep
                        self._append_file_bytes(file, new_data)
                    else:
                        self._append_file_bytes(file, '\n'.encode('utf-8', errors='surrogateescape') + new_data) # changed from os.linesep
        except FileNotFoundError as e:
            print(f"File '{file}' not found: {e}")
        except OSError as e:
            print(f"OS Error: {e}")
    def read(self, file, address=None):
        if not isinstance(file, str):
            raise TypeError("file must be a string.")
        if address is not None and not isinstance(address, list):
            raise TypeError("address must be a list or none")
        if address is None:
            address = []
        if len(address) != 2:
            raise ValueError("address must contain containerid and sectorid")
        containerid, sectorid = address
        if not isinstance(containerid, int) or not isinstance(sectorid, int):
            raise TypeError("containerid and sectorid must be integers")
        if containerid < 0 or sectorid < 0:
            raise ValueError("containerid and sectorid must be >= 0")
        try:
            r = self._read_file_bytes(file).decode('utf-8', errors='surrogateescape')
            pattern = rf'(?s)!<\[{containerid},{sectorid}],(.*?)>!'
            match = re.search(pattern, r)
            if match:
                return match.group(1)
            else:
                return ""
        except FileNotFoundError as e:
            print(f"File '{file}' not found: {e}")
            return ""
        except OSError as e:
            print(f"OS Error: {e}")
            return ""
    def edit(self, file, data, address=None):
        if not isinstance(file, str):
            raise TypeError("file must be a string.")
        if not isinstance(data, str):
            raise TypeError("data must be a string.")
        if address is not None and not isinstance(address, list):
            raise TypeError("address must be a list or None")
        if address is None:
            address = []
        if len(address) != 2:
            raise ValueError("address must contain containerid and sectorid.")
        containerid, sectorid = address
        if not isinstance(containerid, int) or not isinstance(sectorid, int):
            raise TypeError("containerid and sectorid must be integers.")
        if containerid < 0 or sectorid < 0:
            raise ValueError("containerid and sectorid must be >=0.")
        try:
            r = self._read_file_bytes(file).decode('utf-8', errors='surrogateescape')
            if self.check(file, 'sector', [containerid, sectorid]):
                pattern = rf'(!<\[{containerid},{sectorid}],)([^>!]*)>!'
                replacement = rf'\g{data}>!'
                new_content = re.sub(pattern, replacement, r, count=1).encode('utf-8', errors='surrogateescape')
                self._write_file_bytes(file, new_content)
        except FileNotFoundError as e:
            print(f"File '{file}' not found: {e}")
        except OSError as e:
            print(f"OS Error: {e}")
    def change_name(self, file, new_name, containerid):
        if not isinstance(file, str):
            raise TypeError("file must be a string.")
        if not isinstance(new_name, str):
            raise TypeError("new_name  must be a string.")
        if not isinstance(containerid, int):
            raise TypeError("containerid must be an integer.")
        if containerid < 0:
            raise ValueError("Container ID must be >= 0")
        if not new_name:
            raise ValueError("new_name cannot be empty")
        filepath = self._normalize_path(file)
        if not self._check_file_permissions(filepath):
            raise OSError(f"Insufficient permissions to write to '{filepath}'")
        try:
            file_content_bytes = self._read_file_bytes(filepath)
            file_content_str = file_content_bytes.decode('utf-8', errors='surrogateescape')
            file_content_str = file_content_str.replace('\r\n','\n') # Normalize line endings
            pattern = rf"&<{containerid}\^.*?\n" # changed from os.linesep
            replacement = f"&<{containerid}^{new_name}>\n" # changed from os.linesep
            updated_content = re.sub(pattern, replacement, file_content_str).encode('utf-8', errors='surrogateescape')
            lines = updated_content.splitlines()
            non_empty_lines = [line for line in lines if line.strip()]
            final_content = '\n'.encode('utf-8', errors='surrogateescape').join(non_empty_lines) # changed from os.linesep
            self._write_file_bytes(filepath, final_content)
        except OSError as e:
            raise OSError(f"OS Error: {e}")
        except TypeError as e:
            raise TypeError(f"Type Error: {e}")
        except ValueError as e:
            raise ValueError(f"Value Error: {e}")
        except IndexError as e:
            raise IndexError(f"Index Error: {e}")
    def readsectors(self, file, containerid):
        if not isinstance(file, str):
            raise TypeError("file must be a string.")
        if not isinstance(containerid, int):
            raise TypeError("containerid must be an integer.")
        if containerid < 0:
            raise ValueError("containerid must be >= 0")
        try:
            r = self._read_file_bytes(file).decode('utf-8', errors='surrogateescape')
            r = r.replace('\r\n', '\n')  # Normalize line endings
            pattern = rf'(?s)!<\[{containerid},(\d+)],(.*?)>!'
            matches = re.finditer(pattern, r)
            section_data = []
            for match in matches:
                section_id = int(match.group(1))
                value = match.group(2)
                section_data.append((section_id, value))
            section_data.sort(key=lambda x: x[0])
            data = [value for _, value in section_data]
            return data
        except FileNotFoundError as e:
            print(f"File '{file}' not found: {e}")
            return []
        except OSError as e:
            print(f"OS Error: {e}")
            return []
    def numbercontainers(self, file, plogic=False):
        if not isinstance(file, str):
            raise TypeError("file must be a string.")
        if not isinstance(plogic, bool):
            raise TypeError("plogic must be a bool.")
        try:
            r = self._read_file_bytes(file).decode('utf-8', errors='surrogateescape')
            r = r.replace('\r\n', '\n')  # Normalize line endings
            count = len(re.findall(r'\$<(\d+),', r))
            return count if not plogic else count - 1
        except FileNotFoundError as e:
            print(f"File '{file}' not found: {e}")
            return 0
        except OSError as e:
            print(f"OS Error: {e}")
            return 0
    def numbersectors(self, file, containerid, plogic=False):
        if not isinstance(file, str):
            raise TypeError("file must be a string.")
        if not isinstance(containerid, int):
            raise TypeError("containerid must be an integer.")
        if not isinstance(plogic, bool):
            raise TypeError("plogic must be a bool.")
        if containerid < 0:
            raise ValueError("containerid must be >= 0")
        try:
            r = self._read_file_bytes(file).decode('utf-8', errors='surrogateescape')
            r = r.replace('\r\n', '\n')  # Normalize line endings
            pattern = rf'!<\[{containerid},\d+]'
            count = len(re.findall(pattern, r))
            return count if not plogic else count - 1
        except FileNotFoundError as e:
            print(f"File '{file}' not found: {e}")
            return 0
        except OSError as e:
            print(f"OS Error: {e}")
            return 0
    def delete(self, file, address=None):
        if not isinstance(file, str):
            raise TypeError("file must be a string.")
        if address is not None and not isinstance(address, list):
            raise TypeError("address must be a list or none")
        if address is None:
            address = []
        if len(address) != 2:
            raise ValueError("address must contain containerid and sectorid.")
        containerid, sectorid = address
        if not isinstance(containerid, int) or not isinstance(sectorid, int):
            raise TypeError("containerid and sectorid must be integers.")
        if containerid < 0 or sectorid < 0:
            raise ValueError("containerid and sectorid must be >= 0.")
        filepath = self._normalize_path(file)
        if not self._check_file_permissions(filepath):
            raise OSError(f"Insufficient permissions to write to '{filepath}'")
        try:
            file_content_bytes = self._read_file_bytes(filepath)
            file_content_str = file_content_bytes.decode('utf-8', errors='surrogateescape')
            file_content_str = file_content_str.replace('\r\n', '\n')  # normalize line endings.
            pattern = rf'(?s)!<\[{containerid},{sectorid}],(.*?)>!'
            updated_content = re.sub(pattern, '', file_content_str)
            lines = updated_content.split('\n')  # changed from os.linesep
            non_empty_lines = [line for line in lines if line.strip() != '']
            final_content = '\n'.join(non_empty_lines)  # changed from os.linesep
            self._write_file_bytes(filepath, final_content.encode('utf-8', errors='surrogateescape'))
        except FileNotFoundError as e:
            print(f"File '{file}' not found: {e}")
        except OSError as e:
            print(f"OS Error: {e}")
    def drop(self, file, containerid):
        if not isinstance(file, str):
            raise TypeError("file must be a string.")
        if not isinstance(containerid, int):
            raise TypeError("containerid must be an integer.")
        if containerid < 0:
            raise ValueError("containerid must be >= 0")
        filepath = self._normalize_path(file)
        if not self._check_file_permissions(filepath):
            raise OSError(f"Insufficient permissions to write to '{filepath}'")
        try:
            file_content_bytes = self._read_file_bytes(filepath)
            file_content_str = file_content_bytes.decode('utf-8', errors='surrogateescape')
            file_content_str = file_content_str.replace('\r\n', '\n')  # Normalize line endings
            pattern = rf"(?s)\$<{containerid},[^>]*>\n(?:!<\[{containerid},\d+],(?:(?!>!).)*?>!(?:\n(?!<{containerid},).*)?)*"
            updated_content = re.sub(pattern, '', file_content_str, flags=re.DOTALL)
            lines = updated_content.split('\n')
            non_empty_lines = [line for line in lines if line.strip() != '']
            final_content = '\n'.join(non_empty_lines)
            self._write_file_bytes(filepath, final_content.encode('utf-8', errors='surrogateescape'))
        except FileNotFoundError as e:
            print(f"File '{file}' not found: {e}")
        except OSError as e:
            print(f"OS Error: {e}")
    def check(self, file, type, address):
        if not isinstance(file, str):
            raise TypeError("file must be a string.")
        if not isinstance(type, str):
            raise TypeError("type must be a string.")
        if not isinstance(address, list):
            raise TypeError("address must be a list.")
        if type not in ('container', 'sector'):
            raise ValueError("type must be 'container' or 'sector'")
        if type == 'container':
            if len(address) != 2:
                raise ValueError("address for container check must contain containerid and name")
            containerid, name = address
            if not isinstance(containerid, int):
                raise TypeError("containerid must be an integer.")
            if not isinstance(name, str):
                raise TypeError("name must be a string.")
            if containerid < 0:
                raise ValueError("containerid must be >= 0")
            try:
                r = self._read_file_bytes(file).decode('utf-8', errors='surrogateescape')
                r = r.replace('\r\n', '\n')  # Normalize line endings
                pattern = rf'\$<{containerid},{re.escape(name)}>'
                return bool(re.search(pattern, r))
            except FileNotFoundError:
                return False
            except OSError:
                return False
        elif type == 'sector':
            if len(address) != 2:
                raise ValueError("address for sector check must contain containerid and sectorid")
            containerid, sectorid = address
            if not isinstance(containerid, int) or not isinstance(sectorid, int):
                raise TypeError("containerid and sectorid must be integers.")
            if containerid < 0 or sectorid < 0:
                raise ValueError("containerid and sectorid must be >= 0.")
            try:
                r = self._read_file_bytes(file).decode('utf-8', errors='surrogateescape')
                r = r.replace('\r\n', '\n')  # Normalize line endings
                pattern = rf'!<\[{containerid},{sectorid}]'
                return bool(re.search(pattern, r))
            except FileNotFoundError:
                return False
            except OSError:
                return False
        return False
container_data = container_data_class()
class table_data_class:
    def __init__(self):
        super().__init__()
    def _validate_dbfile(self, dbfile):
        if not isinstance(dbfile, str):
            raise TypeError("dbfile must be a string.")
        dbfile_path = self._normalize_path(dbfile)
        if not os.path.exists(dbfile_path):
            raise FileNotFoundError(f"Database file '{dbfile_path}' not found.")
        return dbfile_path
    def _normalize_path(self, file_path, C: bool = False):
        if not isinstance(file_path, str):
            raise TypeError("file_path must be a string.")
        filepath = os.path.abspath(os.path.normpath(file_path))
        if not os.path.exists(filepath):
            if C:
                return filepath
            else:
                raise FileNotFoundError(f"File '{filepath}' does not exist.")
        return filepath
    def _read_file_bytes(self, file_path):
        filepath = self._normalize_path(file_path)
        if not self._check_file_permissions(filepath, stat.S_IRUSR):
            raise PermissionError(f"Insufficient permissions to read file: '{filepath}'")
        try:
            with open(filepath, 'rb') as f:
                return f.read()
        except OSError as e:
            raise OSError(f"OS Error while reading file '{filepath}': {e}")
    def _write_file_bytes(self, file_path, data):
        filepath = self._normalize_path(file_path)
        if not isinstance(data, bytes):
            raise TypeError("data must be bytes.")
        if not self._check_file_permissions(filepath, stat.S_IWUSR):
            self._set_file_permissions(filepath)
            if not self._check_file_permissions(filepath, stat.S_IWUSR):
                raise PermissionError(f"Insufficient permissions to write to file: '{filepath}'")
        try:
            with open(filepath, 'wb') as f:
                f.write(data)
        except OSError as e:
            raise OSError(f"OS Error while writing file '{filepath}': {e}")
    def _append_file_bytes(self, file_path, data):
        filepath = self._normalize_path(file_path)
        if not isinstance(data, bytes):
            raise TypeError("data must be bytes.")
        if not self._check_file_permissions(filepath, stat.S_IWUSR):
            self._set_file_permissions(filepath)
            if not self._check_file_permissions(filepath, stat.S_IWUSR):
                raise PermissionError(f"Insufficient permissions to append to file: '{filepath}'")
        try:
            with open(filepath, 'ab') as f:
                f.write(data)
        except OSError as e:
            raise OSError(f"OS Error while appending file '{filepath}': {e}")
    def _check_file_permissions(self, file_path, required_mode=stat.S_IRUSR | stat.S_IWUSR):
        if not isinstance(file_path, str):
            raise TypeError("file_path must be a string.")
        if not isinstance(required_mode, int):
            raise TypeError("required_mode must be an integer.")
        filepath = self._normalize_path(file_path)
        try:
            mode = stat.S_IMODE(os.stat(filepath).st_mode)
            return bool(mode & required_mode)
        except OSError as e:
            raise OSError(f"OS Error while checking permissions for '{filepath}': {e}")
    def _set_file_permissions(self, file_path, mode=0o660):
        if not isinstance(file_path, str):
            raise TypeError("file_path must be a string.")
        if not isinstance(mode, int):
            raise TypeError("mode must be an integer.")
        filepath = self._normalize_path(file_path)
        try:
            os.chmod(filepath, mode)
        except OSError as e:
            raise OSError(f"OS Error while setting permissions for '{filepath}': {e}")
    def getname(self, file_path, table_id):
        if not isinstance(file_path, str):
            raise TypeError("file_path must be a string.")
        if not isinstance(table_id, int):
            raise TypeError("table_id must be an integer.")
        if table_id < 0:
            raise ValueError("table_id must be >= 0.")
        filepath = self._normalize_path(file_path)
        if not self._check_file_permissions(filepath, stat.S_IRUSR):
            raise PermissionError(f"Insufficient permissions to read file: '{filepath}'")
        try:
            file_content_bytes = self._read_file_bytes(filepath)
            file_content_str = file_content_bytes.decode('utf-8', errors='surrogateescape')
            file_content_str = file_content_str.replace('\r\n', '\n') # Normalize line endings
            pattern = rf'&<{table_id}\^([^>]*)>'
            match = re.search(pattern, file_content_str)
            if match:
                return match.group(1)
            return None
        except OSError as e:
            return None
    def getid(self, file_path, name, plogic=True):
        if not isinstance(file_path, str):
            raise TypeError("file_path must be a string.")
        if not isinstance(name, str):
            raise TypeError("name must be a string.")
        if not isinstance(plogic, bool):
            raise TypeError("plogic must be a bool.")
        if not name:
            raise ValueError("name cannot be empty")
        filepath = self._normalize_path(file_path)
        if not self._check_file_permissions(filepath, stat.S_IRUSR):
            raise PermissionError(f"Insufficient permissions to read file: '{filepath}'")
        try:
            file_content_bytes = self._read_file_bytes(filepath)
            file_content_str = file_content_bytes.decode('utf-8', errors='surrogateescape')
            file_content_str = file_content_str.replace('\r\n', '\n') # Normalize line endings
            pattern = rf'&<(\d+)\^{re.escape(name)}>'
            match = re.search(pattern, file_content_str)
            if match:
                return int(match.group(1)) if plogic else int(match.group(1)) + 1
            return -1 if plogic else 0
        except OSError as e:
            return -1 if plogic else 0
    def numbertables(self, file_path, plogic=False):
        if not isinstance(file_path, str):
            raise TypeError("file_path must be a string.")
        if not isinstance(plogic, bool):
            raise TypeError("plogic must be a bool.")
        filepath = self._normalize_path(file_path)
        if not self._check_file_permissions(filepath, stat.S_IRUSR):
            raise PermissionError(f"Insufficient permissions to read file: '{filepath}'")
        try:
            file_content_bytes = self._read_file_bytes(filepath)
            file_content_str = file_content_bytes.decode('utf-8', errors='surrogateescape')
            file_content_str = file_content_str.replace('\r\n', '\n') # Normalize line endings
            count = len(re.findall(r'&<\d+\^', file_content_str))
            return count if not plogic else count - 1
        except OSError as e:
            return -1 if plogic else 0
    def hcolumn(self, file: str, tableid: int, plogic: bool = False, sprow: int = -1):
        try:
            rawp = os.path.abspath(os.path.normpath(file))
            f = open(rawp, 'rb')
            rawd = f.read()
            f.close()
            r = rawd.decode('utf-8', errors='surrogateescape')
        except (FileNotFoundError,OSError,PermissionError):
            return -1 if plogic else 0
        if sprow == -1:
            pattern = rf'~<\[{tableid};(\d+)\?'
        else:
            pattern = rf'~<\[{tableid};(\d+)\?{sprow}\]'
        matches = re.findall(pattern, r)
        if not matches:
            return -1 if plogic else 0
        try:
            max_col = max(map(int, matches))
            return max_col if plogic else max_col + 1
        except ValueError:
            return -1 if plogic else 0
    def hrow(self, file: str, tableid: int, plogic: bool = False, sprow: int = -1):
        try:
            rawp = os.path.abspath(os.path.normpath(file))
            f = open(rawp, 'rb')
            rawd = f.read()
            f.close()
            r = rawd.decode('utf-8', errors='surrogateescape')
        except (FileNotFoundError, OSError, PermissionError):
            return -1 if plogic else 0
        if sprow == -1:
            pattern = rf'~<\[{tableid};\d+\?(\d+)]'
        else:
            pattern = rf'~<\[{tableid};{sprow}\?(\d+)]' #corrected line.
        matches = re.findall(pattern, r)
        if not matches:
            return -1 if plogic else 0
        try:
            max_row = max(map(int, matches))
            return max_row if plogic else max_row + 1
        except ValueError:
            return -1 if plogic else 0
    def numbercolumns(self, file_path, address=None, plogic=False):
        if not isinstance(file_path, str):
            raise TypeError("file_path must be a string.")
        if address is not None and not isinstance(address, list):
            raise TypeError("address must be a list or None.")
        if address is not None:
            if len(address) != 2:
                raise ValueError("address list must have length 2: (table_id, row_id)")
        if not isinstance(plogic, bool):
            raise TypeError("plogic must be a bool.")
        filepath = self._normalize_path(file_path)
        if not self._check_file_permissions(filepath, stat.S_IRUSR):
            raise PermissionError(f"Insufficient permissions to read file: '{filepath}'")
        try:
            if address:
                return self.hcolumn(filepath, address[0], plogic, address[1])
            return -1 if plogic else 0
        except OSError as e:
            return -1 if plogic else 0
    def numberrows(self, file_path, address=None, plogic=False):
        if not isinstance(file_path, str):
            raise TypeError("file_path must be a string.")
        if address is not None and not isinstance(address, list):
            raise TypeError("address must be a list or None.")
        if address is not None:
            if len(address) != 2:
                raise ValueError("address list must have length 2: [table_id, row_id]")
        if not isinstance(plogic, bool):
            raise TypeError("plogic must be a bool.")
        filepath = self._normalize_path(file_path)
        if not self._check_file_permissions(filepath, stat.S_IRUSR):
            raise PermissionError(f"Insufficient permissions to read file: '{filepath}'")
        try:
            if address:
                return self.hrow(filepath, address[0], plogic, address[1])
            return -1 if plogic else 0
        except OSError as e:
            return -1 if plogic else 0
    def totalcolumns(self, file: str, tableid: int, plogic: bool = False):
        if not isinstance(file, str):
            raise TypeError("file must be a string")
        if not isinstance(tableid, int):
            raise TypeError("tableid must be an integer")
        if not isinstance(plogic, bool):
            raise TypeError("plogic must be a bool")
        filepath = self._normalize_path(file)
        if not self._check_file_permissions(filepath, stat.S_IRUSR):
            raise PermissionError(f"Insufficient permissions to read file: '{filepath}'")
        try:
            return self.hcolumn(filepath, tableid, plogic)
        except OSError as e:
            return -1 if plogic else 0
    def totalrows(self, file: str, tableid: int, plogic: bool = False):
        if not isinstance(file, str):
            raise TypeError("file must be a string")
        if not isinstance(tableid, int):
            raise TypeError("tableid must be an integer")
        if not isinstance(plogic, bool):
            raise TypeError("plogic must be a bool.")
        filepath = self._normalize_path(file)
        if not self._check_file_permissions(filepath, stat.S_IRUSR):
            raise PermissionError(f"Insufficient permissions to read file: '{filepath}'")
        try:
            return self.hrow(filepath, tableid, plogic)
        except OSError as e:
            return -1 if plogic else 0
    def totaltable(self, file: str, tableid: int, plogic: bool = False):
        if not isinstance(file, str):
            raise TypeError("file must be a string")
        if not isinstance(tableid, int):
            raise TypeError("tableid must be an integer")
        if not isinstance(plogic, bool):
            raise TypeError("plogic must be a bool.")
        filepath = self._normalize_path(file)
        if not self._check_file_permissions(filepath, stat.S_IRUSR):
            raise PermissionError(f"Insufficient permissions to read file: '{filepath}'")
        try:
            return [self.hcolumn(filepath, tableid, plogic), self.hrow(filepath, tableid, plogic)]
        except OSError as e:
            return [-1 if plogic else 0, -1 if plogic else 0]
    def insert(self, file: str, data: str, address=None, showmatrix: bool = False):
        if not isinstance(file, str):
            raise TypeError("file must be a string")
        if not isinstance(data, str):
            raise TypeError("data must be a string")
        filepath = self._normalize_path(file)
        if address is None:
            address = []
        if not isinstance(address, list):
            raise TypeError("address must be a list")
        if len(address) != 3:
            raise ValueError("address must contain tableid, columnid, and rowid")
        tableid = address[0]
        columnid = address[1]
        rowid = address[2]
        if not isinstance(tableid, int) or not isinstance(columnid, int) or not isinstance(rowid, int):
            raise TypeError("tableid, columnid, and rowid must be integers")
        if not self._check_file_permissions(filepath, stat.S_IRUSR | stat.S_IWUSR):
            raise PermissionError(f"Insufficient permissions to read and write file: '{filepath}'")
        try:
            file_content_bytes = self._read_file_bytes(filepath)
            file_content_str = file_content_bytes.decode('utf-8', errors='surrogateescape')
            file_content_str = file_content_str.replace('\r\n', '\n')  # Normalize line endings
            r = file_content_str
            info = self.totaltable(filepath, tableid)
            if showmatrix:
                print(columnid, info[0])
                print(rowid, info[1])
            pattern = rf"(?s)~<\[{tableid};{columnid}\?{rowid}],(.*?)>~"
            if not re.search(pattern, r):
                if columnid <= info[0] and rowid <= info[1]:
                    self._append_file_bytes(filepath,
                                            f"{'\n' if r and not r.endswith('\n') else ''}~<[{tableid};{columnid}?{rowid}],{data}>~".encode(
                                                'utf-8', errors='surrogateescape'))  # changed os.linesep
        except OSError as e:
            raise OSError(f"OS Error: {e}")
        except TypeError as e:
            raise TypeError(f"Type Error: {e}")
        except ValueError as e:
            raise ValueError(f"Value Error: {e}")
        except IndexError as e:
            raise IndexError(f"Index Error: {e}")
    def read(self, file: str, address=None):
        if not isinstance(file, str):
            raise TypeError("file must be a string")
        filepath = self._normalize_path(file)
        if address is None:
            address = []
        if not isinstance(address, list):
            raise TypeError("address must be a list")
        if len(address) != 3:
            raise ValueError("address must contain tableid, columnid, and rowid")
        tableid = address[0]
        columnid = address[1]
        rowid = address[2]
        if not isinstance(tableid, int) or not isinstance(columnid, int) or not isinstance(rowid, int):
            raise TypeError("tableid, columnid, and rowid must be integers")
        if not self._check_file_permissions(filepath, stat.S_IRUSR):
            raise PermissionError(f"Insufficient permissions to read file: '{filepath}'")
        try:
            file_content_bytes = self._read_file_bytes(filepath)
            file_content_str = file_content_bytes.decode('utf-8', errors='surrogateescape')
            file_content_str = file_content_str.replace('\r\n', '\n')  # Normalize line endings
            r = file_content_str
            pattern = rf"(?s)~<\[{tableid};{columnid}\?{rowid}],(.*?)>~"
            match = re.search(pattern, r)
            if match:
                return match.group(1)
            return ""
        except OSError as e:
            return ""
        except TypeError as e:
            return ""
        except ValueError as e:
            return ""
        except IndexError as e:
            return ""
    def readcolumns(self, file: str, address=None):
        if not isinstance(file, str):
            raise TypeError("file must be a string")
        filepath = self._normalize_path(file)
        if address is None:
            address = []
        if not isinstance(address, list):
            raise TypeError("address must be a list")
        if len(address) != 2:
            raise ValueError("address must contain tableid and rowid")
        tableid = address[0]
        rowid = address[1]
        if not isinstance(tableid, int) or not isinstance(rowid, int):
            raise TypeError("tableid and rowid must be integers")
        if not self._check_file_permissions(filepath, stat.S_IRUSR):
            raise PermissionError(f"Insufficient permissions to read file: '{filepath}'")
        try:
            file_content_bytes = self._read_file_bytes(filepath)
            file_content_str = file_content_bytes.decode('utf-8', errors='surrogateescape')
            file_content_str = file_content_str.replace('\r\n', '\n')  # Normalize line endings
            r = file_content_str
            pattern = rf"(?s)~<\[{tableid};(\d+)\?{rowid}],(.*?)>~"
            matches = re.finditer(pattern, r)
            column_data = []
            for match in matches:
                column_id = int(match.group(1))
                value = match.group(2)
                column_data.append((column_id, value))
            column_data.sort(key=lambda x: x[0])
            data = [value for _, value in column_data]
            return data
        except OSError as e:
            return []
        except TypeError as e:
            return []
        except ValueError as e:
            return []
        except IndexError as e:
            return []
    def readrows(self, file: str, address=None):
        if not isinstance(file, str):
            raise TypeError("file must be a string")
        filepath = self._normalize_path(file)
        if address is None:
            address = []
        if not isinstance(address, list):
            raise TypeError("address must be a list")
        if len(address) != 2:
            raise ValueError("address must contain tableid and columnid")
        tableid = address[0]
        columnid = address[1]
        if not isinstance(tableid, int) or not isinstance(columnid, int):
            raise TypeError("tableid and columnid must be integers")
        if not self._check_file_permissions(filepath, stat.S_IRUSR):
            raise PermissionError(f"Insufficient permissions to read file: '{filepath}'")
        try:
            file_content_bytes = self._read_file_bytes(filepath)
            file_content_str = file_content_bytes.decode('utf-8', errors='surrogateescape')
            file_content_str = file_content_str.replace('\r\n', '\n')  # Normalize line endings
            r = file_content_str
            row_data = []
            pattern = rf"(?s)~<\[{tableid};{columnid}\?(\d+)],(.*?)>~"
            for match in re.finditer(pattern, r):
                row_id = int(match.group(1))
                value = match.group(2)
                row_data.append((row_id, value))
            row_data.sort(key=lambda x: x[0])
            data = [value for _, value in row_data]
            return data
        except OSError as e:
            return []
        except TypeError as e:
            return []
        except ValueError as e:
            return []
        except IndexError as e:
            return []
    def edit(self, file: str, data: str, address=None):
        if not isinstance(file, str):
            raise TypeError("file must be a string")
        if not isinstance(data, str):
            raise TypeError("data must be a string")
        filepath = self._normalize_path(file)
        if address is None:
            address = []
        if not isinstance(address, list):
            raise TypeError("address must be a list")
        if len(address) != 3:
            raise ValueError("address must contain tableid, columnid, and rowid")
        tableid = address[0]
        columnid = address[1]
        rowid = address[2]
        if not isinstance(tableid, int) or not isinstance(columnid, int) or not isinstance(rowid, int):
            raise TypeError("tableid, columnid, and rowid must be integers")
        if not self._check_file_permissions(filepath, stat.S_IRUSR | stat.S_IWUSR):
            raise PermissionError(f"Insufficient permissions to read and write file: '{filepath}'")
        try:
            file_content_bytes = self._read_file_bytes(filepath)
            file_content_str = file_content_bytes.decode('utf-8', errors='surrogateescape')
            file_content_str = file_content_str.replace('\r\n', '\n')  # Normalize line endings
            pattern = rf"(?s)~<\[{tableid};{columnid}\?{rowid}],(.*?)>~"
            replacement = rf"~<[{tableid};{columnid}?{rowid}],{data}>~"
            updated_content = re.sub(pattern, replacement, file_content_str)
            self._write_file_bytes(filepath, updated_content.encode('utf-8', errors='surrogateescape'))
        except OSError as e:
            raise OSError(f"OS Error: {e}")
        except TypeError as e:
            raise TypeError(f"Type Error: {e}")
        except ValueError as e:
            raise ValueError(f"Value Error: {e}")
        except IndexError as e:
            raise IndexError(f"Index Error: {e}")
    def change_name(self, file: str, new_name: str, tableid: int):
        if not isinstance(file, str):
            raise TypeError("file must be a string")
        if not isinstance(new_name, str):
            raise TypeError("new_name must be a string")
        if not isinstance(tableid, int):
            raise TypeError("tableid must be an integer")
        filepath = self._normalize_path(file)
        if not self._check_file_permissions(filepath, stat.S_IRUSR | stat.S_IWUSR):
            raise PermissionError(f"Insufficient permissions to read and write file: '{filepath}'")
        try:
            file_content_bytes = self._read_file_bytes(filepath)
            file_content_str = file_content_bytes.decode('utf-8', errors='surrogateescape')
            file_content_str = file_content_str.replace('\r\n', '\n')  # Normalize line endings
            pattern = rf"&<{tableid}\^.*?\n"  # changed from os.linesep
            replacement = f"&<{tableid}^{new_name}>\n"  # changed from os.linesep
            updated_content = re.sub(pattern, replacement, file_content_str)
            lines = updated_content.splitlines()
            non_empty_lines = [line for line in lines if line.strip()]
            final_content = '\n'.join(non_empty_lines)  # changed from os.linesep
            self._write_file_bytes(filepath, final_content.encode('utf-8', errors='surrogateescape'))
        except OSError as e:
            raise OSError(f"OS Error: {e}")
        except TypeError as e:
            raise TypeError(f"Type Error: {e}")
        except ValueError as e:
            raise ValueError(f"Value Error: {e}")
        except IndexError as e:
            raise IndexError(f"Index Error: {e}")
    def all_addresses_grouping(self, file: str, tableid: int, filtermode: int):
        # filtermode must be either 0(columns) or 1(rows)
        if not isinstance(file, str):
            raise TypeError("file must be a string")
        if not isinstance(tableid, int):
            raise TypeError("tableid must be an integer")
        if not isinstance(filtermode, int):
            raise TypeError("filtermode must be an integer")
        filepath = self._normalize_path(file)
        if not self._check_file_permissions(filepath, stat.S_IRUSR):
            raise PermissionError(f"Insufficient permissions to read file: '{filepath}'")
        try:
            file_content_bytes = self._read_file_bytes(filepath)
            file_content_str = file_content_bytes.decode('utf-8', errors='surrogateescape')
            file_content_str = file_content_str.replace('\r\n', '\n')  # Normalize line endings
            pattern = rf"~<\[{tableid};(\d+)\?(\d+)]"
            matches = re.findall(pattern, file_content_str)
            addresses = [[tableid, int(col), int(row)] for col, row in matches]
            addresses.sort(key=lambda address: (address[1], address[2]))
            if filtermode < 2:
                grouped = {}
                index = filtermode + 1  # 1 for col, 2 for row
                for address in addresses:
                    key = address[index]
                    if key not in grouped:
                        grouped[key] = []
                    grouped[key].append(address)
                return list(grouped.values())
            else:
                return []
        except OSError as e:
            return []
        except TypeError as e:
            return []
        except ValueError as e:
            return []
        except IndexError as e:
            return []
    def all_addresses_list(self, file: str, tableid: int, totalnum: bool = False):
        if not isinstance(file, str):
            raise TypeError("file must be a string")
        if not isinstance(tableid, int):
            raise TypeError("tableid must be an integer")
        if not isinstance(totalnum, bool):
            raise TypeError("totalnum must be a boolean")
        filepath = self._normalize_path(file)
        if not self._check_file_permissions(filepath, stat.S_IRUSR):
            raise PermissionError(f"Insufficient permissions to read file: '{filepath}'")
        try:
            file_content_bytes = self._read_file_bytes(filepath)
            file_content_str = file_content_bytes.decode('utf-8', errors='surrogateescape')
            file_content_str = file_content_str.replace('\r\n', '\n')  # Normalize line endings
            pattern = rf"~<\[{tableid};(\d+)\?(\d+)]"
            matches = re.findall(pattern, file_content_str)
            addresses = [[tableid, int(col), int(row)] for col, row in matches]
            addresses.sort(key=lambda address: (address[1], address[2]))
            return len(addresses) if totalnum else addresses
        except OSError as e:
            return 0 if totalnum else []
        except TypeError as e:
            return 0 if totalnum else []
        except ValueError as e:
            return 0 if totalnum else []
        except IndexError as e:
            return 0 if totalnum else []
    def delete(self, file: str, address=None):
        if not isinstance(file, str):
            raise TypeError("file must be a string")
        filepath = self._normalize_path(file)
        if address is None:
            address = []
        if not isinstance(address, list):
            raise TypeError("address must be a list")
        if len(address) != 3:
            raise ValueError("address must contain tableid, columnid, and rowid")
        tableid = address[0]
        columnid = address[1]
        rowid = address[2]
        if not isinstance(tableid, int) or not isinstance(columnid, int) or not isinstance(rowid, int):
            raise TypeError("tableid, columnid, and rowid must be integers")
        if not self._check_file_permissions(filepath, stat.S_IRUSR | stat.S_IWUSR):
            raise PermissionError(f"Insufficient permissions to read and write file: '{filepath}'")
        try:
            file_content_bytes = self._read_file_bytes(filepath)
            file_content_str = file_content_bytes.decode('utf-8', errors='surrogateescape')
            file_content_str = file_content_str.replace('\r\n', '\n')  # Normalize line endings
            pattern = rf"(?s)~<\[{tableid};{columnid}\?{rowid}],(.*?)>~"
            updated_content = re.sub(pattern, '', file_content_str)
            # Remove empty lines
            lines = updated_content.split('\n')  # changed from os.linesep
            non_empty_lines = [line for line in lines if line.strip() != '']
            final_content = '\n'.join(non_empty_lines)  # changed from os.linesep
            self._write_file_bytes(filepath, final_content.encode('utf-8', errors='surrogateescape'))
        except OSError as e:
            raise OSError(f"OS Error: {e}")
        except TypeError as e:
            raise TypeError(f"Type Error: {e}")
        except ValueError as e:
            raise ValueError(f"Value Error: {e}")
        except IndexError as e:
            raise IndexError(f"Index Error: {e}")
    def drop(self, file: str, tableid: int):
        if not isinstance(file, str):
            raise TypeError("file must be a string")
        if not isinstance(tableid, int):
            raise TypeError("tableid must be an integer")
        filepath = self._normalize_path(file)
        if not self._check_file_permissions(filepath, stat.S_IRUSR | stat.S_IWUSR):
            raise PermissionError(f"Insufficient permissions to read and write file: '{filepath}'")
        try:
            file_content_bytes = self._read_file_bytes(filepath)
            file_content_str = file_content_bytes.decode('utf-8', errors='surrogateescape')
            file_content_str = file_content_str.replace('\r\n', '\n')  # Normalize line endings
            # Corrected Regex for multi-line data
            pattern = rf"&<{tableid}\^.*?\n(?:~<\[{tableid};\d+\?\d+\](?:(?!~<\[).)*?>~(?:\n.*?)*?\n?)*"  # changed from os.linesep
            updated_content = re.sub(pattern, '', file_content_str, flags=re.MULTILINE | re.DOTALL)
            # Remove empty lines that were created by re.sub
            lines = updated_content.split('\n')  # changed from os.linesep
            non_empty_lines = [line for line in lines if line.strip() != '']
            final_content = '\n'.join(non_empty_lines)  # changed from os.linesep
            self._write_file_bytes(filepath, final_content.encode('utf-8', errors='surrogateescape'))
        except OSError as e:
            raise OSError(f"OS Error: {e}")
        except TypeError as e:
            raise TypeError(f"Type Error: {e}")
        except ValueError as e:
            raise ValueError(f"Value Error: {e}")
        except IndexError as e:
            raise IndexError(f"Index Error: {e}")
    def export_tables_to_excel(self, dbfile: str, filepath: str):
        dbfile_path = self._validate_dbfile(dbfile)
        if not isinstance(filepath, str):
            raise TypeError("filepath must be a string.")
        filepath = self._normalize_path(filepath, True)
        if not filepath.lower().endswith(('.xlsx', '.xlsm')):
            filepath = f"{filepath}.xlsx"
        if os.path.isdir(filepath):
            raise OSError(f"'{filepath}' is a directory, not a file.")
        try:
            num_sheets = self.numbertables(dbfile_path, False)
            total_items = 0
            for table_id in range(num_sheets):
                total_items += self.all_addresses_list(dbfile_path, table_id, True)
            data_list = []
            raw_data = []
            stuff_list = []
            for main in range(num_sheets):
                raw_data.append(self.all_addresses_list(dbfile_path, main))
            sraw_data = inner_functions.combine_lists(raw_data)
            for m in range(total_items):
                address = sraw_data[m]
                try:
                    value = self.read(dbfile_path, [address[0], address[1], address[2]])
                    stuff_list.append(value)
                except Exception as e:
                    print(f"Error reading data at {address}: {e}", file=sys.stderr)
                    stuff_list.append(None)
            data_list.append(inner_functions.add_data_to_inner_lists(sraw_data, stuff_list))
            data_list = inner_functions.combine_lists(data_list)
            created_sheets = {}
            if os.path.isfile(filepath):
                workbook = openpyxl.load_workbook(filepath)
            else:
                workbook = openpyxl.Workbook()
                std = workbook['Sheet']
                workbook.remove(std)
            for item in range(len(data_list)):
                if len(data_list[item]) == 4:
                    table_id, col_id_0, row_id_0, value = data_list[item]
                    table_name = self.getname(dbfile_path, table_id)
                    if not isinstance(table_name, str) or not re.match(r'^[a-zA-Z0-9_\- ]+$', table_name):
                        print(f"Warning: Invalid table name '{table_name}'. Skipping data entry.")
                        continue
                    if table_name not in created_sheets:
                        if table_name not in workbook.sheetnames:
                            workbook.create_sheet(table_name)
                        created_sheets[table_name] = True
                    if table_name in workbook.sheetnames:
                        sheet = workbook[table_name]
                        col_id_1 = col_id_0 + 1
                        row_id_1 = row_id_0 + 1
                        sheet.cell(row=row_id_1, column=col_id_1, value=value)
                    else:
                        print(f"Warning: Sheet '{table_name}' not found.")
                else:
                    print(f"Warning: Invalid data item {item}. Expected [table_id, column_id, row_id, value].")
            workbook.save(filepath)
            print(f"Data inserted into '{filepath}'.")
        except (TypeError, FileNotFoundError, OSError) as e:
            print(f"Error: {e}", file=sys.stderr)
            raise
        except Exception as e:
            print(f"An unexpected error occurred: {e}", file=sys.stderr)
            raise
table_data = table_data_class()
class OtherClass:
    def __init__(self):
        pass
    def _normalize_path(self, file_path):
        if not isinstance(file_path, str):
            raise TypeError("file_path must be a string.")
        filepath = os.path.abspath(os.path.normpath(file_path))
        if not os.path.exists(filepath):
            raise FileNotFoundError(f"File '{filepath}' does not exist.")
        return filepath
    def _read_file_bytes(self, file_path):
        filepath = self._normalize_path(file_path)
        if not self._check_file_permissions(filepath, stat.S_IRUSR):
            raise PermissionError(f"Insufficient permissions to read file: '{filepath}'")
        try:
            with open(filepath, 'rb') as f:
                return f.read()
        except OSError as e:
            raise OSError(f"OS Error while reading file '{filepath}': {e}")
    def _write_file_bytes(self, file_path, data):
        filepath = self._normalize_path(file_path)
        if not isinstance(data, bytes):
            raise TypeError("data must be bytes.")
        if not self._check_file_permissions(filepath, stat.S_IWUSR):
            self._set_file_permissions(filepath)
            if not self._check_file_permissions(filepath, stat.S_IWUSR):
                raise PermissionError(f"Insufficient permissions to write to file: '{filepath}'")
        try:
            with open(filepath, 'wb') as f:
                f.write(data)
        except OSError as e:
            raise OSError(f"OS Error while writing file '{filepath}': {e}")
    def _check_file_permissions(self, file_path, required_mode=stat.S_IRUSR | stat.S_IWUSR):
        if not isinstance(file_path, str):
            raise TypeError("file_path must be a string.")
        if not isinstance(required_mode, int):
            raise TypeError("required_mode must be an integer.")
        filepath = self._normalize_path(file_path)
        try:
            mode = stat.S_IMODE(os.stat(filepath).st_mode)
            return bool(mode & required_mode)
        except OSError as e:
            raise OSError(f"OS Error while checking permissions for '{filepath}': {e}")
    def _set_file_permissions(self, file_path, mode=0o660):
        if not isinstance(file_path, str):
            raise TypeError("file_path must be a string.")
        if not isinstance(mode, int):
            raise TypeError("mode must be an integer.")
        filepath = self._normalize_path(file_path)
        try:
            os.chmod(filepath, mode)
        except OSError as e:
            raise OSError(f"OS Error while setting permissions for '{filepath}': {e}")
    def clear(self, file: str):
        if not isinstance(file, str):
            raise TypeError("file must be a string.")
        filepath = self._normalize_path(file)
        try:
            self._write_file_bytes(filepath, b'#POWER_DB')
        except OSError as e:
            raise OSError(f"Error occurred while clearing/writing to file: {e}")
    def check(self, file: str, itemtype: str, address=None):
        if not isinstance(file, str):
            raise TypeError("file must be a string.")
        if not isinstance(itemtype, str):
            raise TypeError("itemtype must be a string.")
        if itemtype.lower() not in ('container', 'table', 'sector', 'cell'):
            raise ValueError("Invalid itemtype.  Must be 'container', 'table', 'sector', or 'cell'.")
        filepath = self._normalize_path(file)
        if not self._check_file_permissions(filepath, stat.S_IRUSR):
            raise PermissionError(f"Insufficient permissions to read file: '{filepath}'")
        try:
            file_content_bytes = self._read_file_bytes(filepath)
            file_content_str = file_content_bytes.decode('utf-8', errors='surrogateescape')
            file_content_str = file_content_str.replace('\r\n', '\n') # Normalize line endings.
        except OSError as e:
            raise OSError(f"Error reading file: {e}")
        if itemtype.lower() in ('container', 'table', 'sector', 'cell'):
            if address is None:
                raise ValueError(f"address is required for itemtype '{itemtype}'.")
            if not isinstance(address, list):
                raise TypeError("address must be a list.")
        if itemtype.lower() == 'container':
            if len(address) != 2:
                raise ValueError("address for container must contain containerid and name.")
            containerid = address[0]
            name = address[1]
            if not isinstance(containerid, int):
                raise TypeError("containerid must be an int")
            if not isinstance(name, str):
                raise TypeError("name must be a str")
            search_string = f'$<{containerid},{name}>'
            return search_string in file_content_str
        elif itemtype.lower() == 'table':
            if len(address) != 2:
                raise ValueError("address for table must contain tableid and name.")
            tableid = address[0]
            name = address[1]
            if not isinstance(tableid, int):
                raise TypeError("tableid must be an int")
            if not isinstance(name, str):
                raise TypeError("name must be a str")
            search_string = f'&<{tableid}^{name}>'
            return search_string in file_content_str
        elif itemtype.lower() == 'sector':
            if len(address) != 2:
                raise ValueError("address for sector must contain containerid and sectorid.")
            containerid = address[0]
            sectorid = address[1]
            if not isinstance(containerid, int):
                raise TypeError("containerid must be an int")
            if not isinstance(sectorid, int):
                raise TypeError("sectorid must be an int")
            search_string = f'!<[{containerid},{sectorid}],'
            return search_string in file_content_str
        elif itemtype.lower() == 'cell':
            if len(address) != 3:
                raise ValueError("address for cell must contain tableid, columnid, and rowid.")
            tableid = address[0]
            columnid = address[1]
            rowid = address[2]
            if not isinstance(tableid, int):
                raise TypeError("tableid must be an int")
            if not isinstance(columnid, int):
                raise TypeError("columnid must be an int")
            if not isinstance(rowid, int):
                raise TypeError("rowid must be an int")
            search_string = f'~<[{tableid};{columnid}?{rowid}],'
            return search_string in file_content_str
        else:
            return False
other = OtherClass()