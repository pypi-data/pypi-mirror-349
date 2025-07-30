# -*- coding: UTF-8 -*-
import os
from copy import copy

from openpyxl import Workbook
from openpyxl.reader.excel import load_workbook
from openpyxl.styles import Font, Alignment, NamedStyle, PatternFill
from openpyxl.utils import get_column_letter


class ExcelOp(object):
    """操作xls 表格工具类"""

    def __init__(self, file):
        self.file = file
        self.wb = load_workbook(self.file)
        self.ws = self.wb[self.wb.sheetnames[0]]  # 默认打开第一个表

    def get_all_sheet(self):
        """获取所有工作表名称"""
        return self.wb.sheetnames

    def get_row_clo_num(self):
        """获取表格的总行数和总列数"""
        rows = self.ws.max_row
        columns = self.ws.max_column
        return rows, columns

    def get_cell_value(self, row, column):
        """获取某个单元格的值"""
        cell_value = self.ws.cell(row=row, column=column).value
        return cell_value

    def get_col_value(self, column):
        """获取某列的所有值"""
        rows = self.ws.max_row
        column_data = []
        for i in range(1, rows + 1):
            cell_value = self.ws.cell(row=i, column=column).value
            column_data.append(cell_value)
        return column_data

    def get_row_value(self, row):
        """获取某行所有值"""
        columns = self.ws.max_column
        row_data = []
        for i in range(1, columns + 1):
            cell_value = self.ws.cell(row=row, column=i).value
            row_data.append(cell_value)
        return row_data

    def set_work_sheet(self, sheet_index=0):
        """设置当前工作表"""
        self.ws = self.wb[self.wb.sheetnames[sheet_index]]

    def set_cell_value(self, row, column, cell_value, commit=True):
        """设置某个单元格的值"""
        self.ws.cell(row=row, column=column).value = cell_value
        if commit:
            self.save_sheet()

    def set_row_value(self, content: "list or tuple", row_: int, commit=True):
        """设置某行所有的值"""
        if not isinstance(content, (list, tuple, str)):
            # 可迭代类型 及 参数无法写入
            raise ValueError("Parameter type error. Current row content cannot be set")

        if isinstance(content, str):
            content = [content, ]

        for i in range(0, len(content)):
            self.ws.cell(row=row_, column=i + 1, value=content[i])
        if commit:
            self.save_sheet()

    def save_sheet(self):
        """保存表变更值"""
        try:
            self.wb.save(self.file)
        except IOError as e:
            print("设置某行所有的值错误: %s" % e)
        finally:
            self.wb.close()

    @staticmethod
    def copy_cell_properties(source_cell, target_cell):
        """复制单元格及属性"""
        target_cell.data_type = copy(source_cell.data_type)
        target_cell.fill = copy(source_cell.fill)
        if source_cell.has_style:
            # 样式
            target_cell._style = copy(source_cell._style)
            target_cell.font = copy(source_cell.font)
            target_cell.border = copy(source_cell.border)
            target_cell.fill = copy(source_cell.fill)
            target_cell.number_format = copy(source_cell.number_format)
            target_cell.protection = copy(source_cell.protection)
            target_cell.alignment = copy(source_cell.alignment)

        if source_cell.hyperlink:
            # 超链接
            target_cell._hyperlink = copy(source_cell.hyperlink)

        if source_cell.comment:
            # 注解
            target_cell.comment = copy(source_cell.comment)


class CreateExcel:
    """CreateExcel"""

    def __init__(self, msg):
        # 创建五个表
        self.workbook = Workbook()
        self.workbook.active.title = "升级版本软件包"
        self.workbook.create_sheet("降低版本软件包")
        self.workbook.create_sheet("同版本软件包")
        self.workbook.create_sheet("删除软件包")
        self.workbook.create_sheet("新增软件包")
        self.workbook.create_sheet("文件目录比对")
        self.workbook.create_sheet("文件内容比对")

        # 填充
        self.blue_fill = PatternFill('solid', fgColor='4F81BD')

        # 顶部样式
        self.header_title = NamedStyle(name="header_title")
        self.header_title.font = Font(name=u"宋体", sz=14, bold=True)
        self.header_title.alignment = Alignment(horizontal='center', vertical="center")
        self.header_title.fill = self.blue_fill

        # 正文样式
        self.header_name = NamedStyle(name="header_name")
        self.header_name.font = Font(name=u"宋体", sz=14)
        self.header_name.alignment = Alignment(vertical="center", wrap_text=True)

        # 添加顶部信息
        for i in range(0, 7):
            self.__add_header__(i)
            self.workbook.active.cell(2, 1).value = msg
        self.workbook.active = 0

    def __add_header_up_down__(self, index):
        self.workbook.active = index
        self.workbook.active.merge_cells('A1:D1')
        self.workbook.active.merge_cells('A2:D2')
        self.workbook.active.merge_cells('A3:B3')
        self.workbook.active.merge_cells('C3:D3')
        self.workbook.active.cell(3, 1, value='SRPM').alignment = Alignment(horizontal='center', vertical='center')
        self.workbook.active.cell(3, 3, value='RPM').alignment = Alignment(horizontal='center', vertical='center')
        self.workbook.active.cell(1, 1).value = self.workbook.sheetnames[index]
        self.workbook.active.cell(1, 1).style = self.header_title
        self.workbook.active.cell(2, 1).style = self.header_name
        for i in range(1, 5):
            if (i % 2) == 0:
                self.workbook.active.cell(4, i, 'B').alignment = Alignment(horizontal='center', vertical='center')
            else:
                self.workbook.active.cell(4, i, 'A').alignment = Alignment(horizontal='center', vertical='center')
            self.workbook.active.column_dimensions[get_column_letter(i)].width = 70
            if i < 3:
                self.workbook.active.row_dimensions[i].height = 40

    def __add__header_add_del__(self, index):
        self.workbook.active = index
        self.workbook.active.merge_cells('A1:B1')
        self.workbook.active.merge_cells('A2:B2')
        self.workbook.active.cell(1, 1).value = self.workbook.sheetnames[index]
        self.workbook.active["A1"].style = self.header_title
        self.workbook.active["A2"].style = self.header_name
        for i in range(1, 3):
            self.workbook.active.column_dimensions[get_column_letter(i)].width = 70
            self.workbook.active.row_dimensions[i].height = 40
        self.workbook.active.cell(3, 1, value='SRPM').alignment = Alignment(horizontal='center', vertical='center')
        self.workbook.active.cell(3, 2, value='RPM').alignment = Alignment(horizontal='center', vertical='center')

    def __add_header_list__(self, index):
        self.workbook.active = index
        self.workbook.active.merge_cells('A1:C1')
        self.workbook.active.merge_cells('A2:C2')
        self.workbook.active.cell(1, 1).value = self.workbook.sheetnames[index]
        self.workbook.active["A1"].style = self.header_title
        self.workbook.active["A2"].style = self.header_name
        for i in range(1, 4):
            if i == 1:
                self.workbook.active.column_dimensions[get_column_letter(i)].width = 50
            else:
                self.workbook.active.column_dimensions[get_column_letter(i)].width = 35
            if i != 3:
                self.workbook.active.row_dimensions[i].height = 40
        self.workbook.active.cell(3, 1, value='文件名').alignment = Alignment(horizontal='center', vertical='center')
        self.workbook.active.cell(3, 2, value='A(md5)').alignment = Alignment(horizontal='center', vertical='center')
        self.workbook.active.cell(3, 3, value='B(md5)').alignment = Alignment(horizontal='center', vertical='center')

    def __add_header_content__(self, index):
        self.workbook.active = index
        self.workbook.active.merge_cells('A1:B1')
        self.workbook.active.merge_cells('A2:B2')
        self.workbook.active.cell(1, 1).value = self.workbook.sheetnames[index]
        self.workbook.active["A1"].style = self.header_title
        self.workbook.active["A2"].style = self.header_name
        self.workbook.active.column_dimensions['A'].width = 30
        self.workbook.active.row_dimensions[1].height = 40
        self.workbook.active.column_dimensions['B'].width = 100
        self.workbook.active.row_dimensions[2].height = 40
        self.workbook.active.cell(3, 1, value='文件名').alignment = Alignment(horizontal='center', vertical='center')
        self.workbook.active.cell(3, 2, value='差异内容').alignment = Alignment(horizontal='center', vertical='center')

    def __add_header__(self, index):
        if index < 0 or index > 7:
            return "参数错误"

        if 0 <= index < 2:
            self.__add_header_up_down__(index)
        elif 2 <= index <= 4:
            self.__add__header_add_del__(index)
        elif index == 5:
            self.__add_header_list__(index)
        else:
            self.__add_header_content__(index)

    def add_compare_info(self, index, info_a, info_b, compare_type):
        """
        添加源码包信息
        @param index: 表
        @param info_a:  添加信息a
        @param info_b: 添加信息b
        @param compare_type: 信息类型
        @return:
        """
        if index < 0 or index >= 2:
            return "参数有误"
        self.workbook.active = index
        if compare_type.lower() == "srpm":
            for i in range(0, len(info_a)):
                self.workbook.active.cell(i + 5, 1).value = info_a[i]
            for i in range(0, len(info_b)):
                self.workbook.active.cell(i + 5, 2).value = info_b[i]
        elif compare_type.lower() == 'rpm':
            for i in range(0, len(info_a)):
                self.workbook.active.cell(i + 5, 3).value = info_a[i]
            for i in range(0, len(info_b)):
                self.workbook.active.cell(i + 5, 4).value = info_b[i]

    def add_common(self, index, info, info_type):
        """
        @param index: 表
        @param info: 添加信息
        @param info_type:  添加信息类型
        @return:
        """
        if index <= 1 or index > 4:
            return "参数有误"
        self.workbook.active = index
        if info_type.lower() == 'srpm':
            for i in range(0, len(info)):
                self.workbook.active.cell(i + 4, 1).value = info[i]
        elif info_type.lower() == 'rpm':
            for i in range(0, len(info)):
                self.workbook.active.cell(i + 4, 2).value = info[i]
        self.workbook.active = 0

    def add_files(self, index, files, a_md5, b_md5):
        if index != 5:
            return "参数有误"
        self.workbook.active = index
        for _ in range(0, len(files)):
            self.workbook.active.cell(_ + 4, 1).value = files[_]
            self.workbook.active.cell(_ + 4, 2).value = a_md5[_]
            self.workbook.active.cell(_ + 4, 3).value = b_md5[_]
        self.workbook.active = 0

    def add_content_diff(self, index, files, contents, initrd_diff):
        if index != 6:
            return "参数有误"
        self.workbook.active = index
        for _ in range(0, len(files) + 1):
            if _ < len(files):
                self.workbook.active.cell(_ + 4, 1).value = files[_]
                self.workbook.active.cell(_ + 4, 2).value = contents[_]
            else:
                self.workbook.active.cell(_ + 4, 1).value = initrd_diff['fp']
                self.workbook.active.cell(_ + 4, 2).value = initrd_diff['diff_content']
            self.workbook.active.cell(_ + 4, 2).alignment = Alignment(wrap_text=True)
        self.workbook.active = 0

    def save(self, save_path):
        """
        保存表
        @param save_path: 保存路径
        @return:
        """
        if os.path.exists(save_path):
            print(save_path + " 此文件会被覆盖。------------")
        self.workbook.save(save_path)
