import os
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from PyQt5.QtWidgets import QMessageBox
import platform

class HBMDataMerge:
    """HBM data merging class"""
    def __init__(self, data_path, output_file, tester, log_callback=None):
        self.data_path = data_path
        self.output_file = f"{output_file}_HBM.txt"
        self.tester = tester
        self.total_files = 0
        self.processed_files = 0

        self.errors = []  # Used to record processing errors

        self.log_callback = log_callback

    def process(self):
        # Make sure the output directory exists
        os.makedirs(os.path.dirname(self.output_file), exist_ok=True)
        
        # Add the actual HBM merging logic here
        print("Processing HBM data...")
        """Main method that performs data processing"""
        try:
            self._log("=== Initialize HBM data processing ===")
            self._log(f"Input Path: {self.data_path}")
            self._log(f"Output File: {self.output_file}")
            self._log(f"Tester selected: {self.tester}")

            with open(self.output_file, 'w', encoding='utf-8') as txt_file:
                excel_files = self._get_excel_files()
                self.total_files = len(excel_files)
                self._log(f"Found {self.total_files} Excel files")

                for idx, (file_path, rel_path) in enumerate(excel_files, 1):
                    self._log(f"Processing File ({idx}/{self.total_files}): {rel_path}")
                    self._process_single_file(file_path, rel_path, txt_file)
                    self.processed_files += 1

            self._log("=== Data processing completed ===")

            # Add completion hint
            message = f"Data merge completed!\nOutput File: {self.output_file}"

            if self.errors:
                self._log("\nErrors encountered during processing: ")
                for error in self.errors:
                    self._log(f"- {error}")

            QMessageBox.information(
                None,
                "Processing completed", 
                message,
                QMessageBox.Ok
            )
        
        except Exception as e:
            error_msg = f"Fatal Error: {str(e)}"
            self.errors.append(error_msg)
            self._log(error_msg)

                # Add after the dialog:
        if platform.system() == "Windows":
            os.startfile(os.path.dirname(self.output_file))
        elif platform.system() == "Darwin":
            os.system(f"open '{os.path.dirname(self.output_file)}'")
        else:
            os.system(f"xdg-open '{os.path.dirname(self.output_file)}'")

    def _log(self, message):
        """Unified log processing method"""
        if self.log_callback:
            self.log_callback(message)
        print(message)  # At the same time, the console output

    def _get_excel_files(self):
        """Recursively get Excel files in all subfolders"""
        excel_files = []
        for root, dirs, files in os.walk(self.data_path):
            for file in files:
                if file.lower().endswith('.xlsx'):
                    full_path = os.path.join(root, file)
                    # Calculating relative paths
                    rel_path = os.path.relpath(full_path, self.data_path)
                    excel_files.append((full_path, rel_path))
        return excel_files
    
    def _process_single_file(self, file_path, rel_path, txt_file):
        """Processing a single Excel file"""
        try:
            wb = load_workbook(file_path, read_only=True)
            
            # Check if the target worksheet exists
            if '14ns to 20ns' not in wb.sheetnames:
                raise ValueError(f"Worksheet '14ns to 20ns' does not exist")
            
            ws = wb['14ns to 20ns']
            
            # Write file header (optimize line breaks)
            #header = [
            #    '=' * 40,
            #    f"Document source: {os.path.basename(rel_path)}",
            #    '=' * 40
            #]
            # txt_file.write('\n'.join(header) + '\n')
            
            # Processing of the G1-H47 region
            self._write_range(ws, 'G1', 'H47', txt_file)
            
            # Process A1 and the last row of E column
            self._write_column_data(ws, txt_file)
            
            wb.close()
            
        except Exception as e:
            error_msg = f"Error processing files {os.path.basename(file_path)} : {str(e)}"
            self.errors.append(error_msg)
            print(error_msg)

    def _write_range(self, ws, start_cell, end_cell, txt_file):
        """Write data to the specified cell range"""
        # The specific cell coordinates that need to be filtered
        filtered_g_rows = {8, 13, 19, 28, 29, 35}

        # Parsing start and end coordinates
        start_row = ws[start_cell].row
        start_col = ws[start_cell].column
        end_row = ws[end_cell].row
        end_col = ws[end_cell].column

        # Iterate over rows in a safe way
        for row_idx in range(start_row, end_row + 1):
            line_parts = []
            for col_idx in range(start_col, end_col + 1):
                # Get the cell object
                cell = ws.cell(row=row_idx, column=col_idx)
                
                # Check if filtering is needed
                if col_idx == 7 and row_idx in filtered_g_rows:  # G column = 7
                    value = ''
                else:
                    try:
                        value = str(cell.value).strip() if cell.value else ''
                    except AttributeError:
                        value = ''
                
                line_parts.append(value)
            
            # Filter empty lines and write
            if any(line_parts):
                txt_file.write('\t'.join(line_parts) + '\n')

        txt_file.write(f"Tester selected: {self.tester}" + '\n')

    def _write_column_data(self, ws, txt_file):
        """Process columns A-E to the last row with data"""
        # Determine the line number of the last line
        last_row = 1
        for row in ws.iter_rows(min_col=1, max_col=1):  # Check Column A
            if row[0].value is None:
                break
            last_row = row[0].row
        
        # Write A1-E{last_row} data
        for row in ws.iter_rows(min_row=1, max_row=last_row,
                              min_col=1, max_col=5):
            line = '\t'.join(str(cell.value) if cell.value is not None else ''
                           for cell in row)
            if line.strip():  # Filter empty lines
                txt_file.write(line + '\n')